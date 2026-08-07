import tempfile
import unittest
import sqlite3
import zipfile
from pathlib import Path
from unittest.mock import patch

import server
from openpyxl import Workbook, load_workbook

DATA_ROOT = server.ROOT.parent / "测试数据"
H209_ROOT = DATA_ROOT / "指套管数据1" / "H209"
N208_ROOT = DATA_ROOT / "指套管数据1" / "N208"


def write_sum(group: Path, unit="2", outage="H209"):
    (group / "SUR000C000I000.SUM").write_text(
        f'<?xml version="1.0" encoding="utf-8"?><Summary><Site><Owner>LHNP</Owner><SiteCode>LHNP</SiteCode><Unit>{unit}</Unit><Component>TH</Component><Outage>{outage}</Outage></Site><Equipment><Tester>TEDDY+A</Tester></Equipment><Probe><Model>TH048NAF25A</Model></Probe></Summary>',
        encoding="utf-8",
    )


def write_ect(group: Path, name: str, row: int, col: int, entry: int, marker=""):
    xml = f'xx<?xml version="1.0" encoding="utf-8"?><HeaderTube><Entry>{entry}</Entry><Row>{row}</Row><Col>{col}</Col><DateTime>2026/1/1 {marker}</DateTime></HeaderTube>yy'
    (group / name).write_bytes(xml.encode("utf-8"))


def report_entry(filename, row, col, entry, uid, indication="WAR", location="P1+10", channel="P1: 4-6", percent="20", calgroup="TH2I09CAL00001", measurement="Point"):
    return f'<ReportEntry><Row>{row}</Row><Column>{col}</Column><Entry>{entry}</Entry><Volts>1.2</Volts><Degrees>0</Degrees><Percent>{percent}</Percent><Indication>{indication}</Indication><Channel>{channel}</Channel><ChannelID>4</ChannelID><Location>{location}</Location><MeasurementType>{measurement}</MeasurementType><Probe>TH048NAF25A</Probe><Analyst>A</Analyst><Analysis>Secondary</Analysis><Datapoint>80</Datapoint><Uid>{uid}</Uid><Filename>{filename}</Filename><Filepath>Z:\\old\\{filename}</Filepath><Calgroup>{calgroup}</Calgroup><Distance>10</Distance><Extent>2</Extent><Length>3</Length><Width>4</Width><Comment>note</Comment></ReportEntry>'


def write_report(group: Path, entries: list[str], name="Report-final.rpt"):
    (group / name).write_text('<?xml version="1.0" encoding="utf-8"?><CitecReport>' + "".join(entries) + "</CitecReport>", encoding="utf-8")


class ParserTests(unittest.TestCase):
    def test_health_status_identifies_expected_local_service(self):
        health = server.health_status()
        self.assertTrue(health["ok"])
        self.assertEqual(health["service"], "thimble-local")
        self.assertEqual(health["version"], server.SERVICE_VERSION)

    def test_position_mappings(self):
        self.assertEqual(server.position_for(1, 1), "L11")
        self.assertEqual(server.position_for(2, 1), "B5")
        self.assertEqual(server.position_for(2, 27), "J15")
        self.assertEqual(server.position_for(1, 15), "M5")
        self.assertEqual(server.position_for(2, 42), "M5")
        self.assertEqual(server.position_for(2, 3), "E11")
        self.assertEqual(server.position_for(2, 47), "L5")
        self.assertEqual(server.position_for(2, 50), "L4")
        self.assertEqual(len({server.position_for(1, tube) for tube in range(1, 51)}), 50)
        self.assertEqual(len({server.position_for(2, tube) for tube in range(1, 51)}), 50)

    def test_database_removes_out_of_range_legacy_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "test.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    connection.execute("INSERT INTO findings(outage,unit_id,thimble_id,imported_at) VALUES('N208',2,82,'now')")
                server.init_db()
                with server.connect() as connection:
                    self.assertEqual(connection.execute("SELECT COUNT(*) FROM findings").fetchone()[0], 0)

    def test_database_deduplication_is_audited_and_preserves_distinct_uid(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "dedup.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    row = ("H209", 2, 1, "B5", 4, "NDD", None, "same.ect", "TH2I09CAL00001", "uid-a", "now")
                    connection.executemany("""INSERT INTO findings(
                        outage,unit_id,thimble_id,position,entry_no,indication,percent,filename,calgroup,uid,imported_at
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""", [row, row, row[:-2] + ("uid-b", "now")])
                server.init_db()
                with server.connect() as connection:
                    self.assertEqual(connection.execute("SELECT COUNT(*) FROM findings").fetchone()[0], 2)
                    self.assertEqual(connection.execute("SELECT COUNT(*) FROM duplicate_records WHERE action='removed'").fetchone()[0], 1)
                summary = server.duplicate_summary()
                self.assertEqual(summary["findings"], 2)
                self.assertEqual(summary["duplicates"], 1)

    def test_core_severity_uses_valid_percent_and_ignores_ndd(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "severity.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    connection.executemany(
                        """INSERT INTO findings(
                            outage,unit_id,thimble_id,position,indication,percent,location,measurement_type,imported_at
                        ) VALUES(?,?,?,?,?,?,?,?,?)""",
                        [
                            ("H209", 2, 42, "M5", "WAR", 22, "P1", "Vmax", "now"),
                            ("H209", 2, 20, "D7", "NDD", 80, "P1+10", "Vmax", "now"),
                            ("H209", 2, 21, "G7", "WAR", None, "P1", "Vmax", "now"),
                            ("H209", 2, 1, "B5", "WAR", 47, "P1", "Vmax", "now"),
                            ("H209", 2, 3, "E11", "", 35, "P4+20", "Vmax", "now"),
                        ],
                    )
                result = server.query_findings({"unit": ["2"], "page": ["1"], "size": ["1"]})
                tube_result = server.query_findings({"unit": ["2"], "thimble": ["1"], "page": ["1"], "size": ["100"]})
            severity = {item["thimble_id"]: item["percent"] for item in result["core_items"]}
            self.assertEqual(severity, {1: 47, 3: 35, 42: 22})
            self.assertEqual(len(result["items"]), 1)
            self.assertEqual(tube_result["total"], 1)
            self.assertEqual(server.display_indication({"indication": "", "location": "P4+20", "measurement_type": "Vmax"}), "未标注")
            self.assertEqual(server.display_indication({"indication": "", "location": "", "measurement_type": "None"}), "NDD")

    def test_site_filter_uses_explicit_site_code(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "site-filter.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    connection.executemany(
                        "INSERT INTO findings(outage,unit_id,thimble_id,site_code,position,imported_at) VALUES(?,?,?,?,?,?)",
                        [("H209", 2, 1, "LHNP", "B5", "now"), ("H209", 2, 2, "DNP", "C8", "now")],
                    )
                result = server.query_findings({"site": ["LHNP"], "page": ["1"], "size": ["50"]})
            self.assertEqual(result["total"], 1)
            self.assertEqual(result["items"][0]["site_code"], "LHNP")

    def test_findings_page_size_supports_500_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "page-size.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    connection.executemany(
                        "INSERT INTO findings(outage,unit_id,thimble_id,position,indication,percent,imported_at) VALUES(?,?,?,?,?,?,?)",
                        [("H209", 2, index % 50 + 1, "B5", "WAR", 20, "now") for index in range(550)],
                    )
                result = server.query_findings({"page": ["1"], "size": ["500"]})
            self.assertEqual(result["size"], 500)
            self.assertEqual(len(result["items"]), 500)
            self.assertEqual(result["pages"], 2)

    def test_word_reports_use_selected_rows_and_comparison_result(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            db = root / "reports.db"
            with patch.object(server, "DB_PATH", db), patch.object(server, "REPORT_DIR", root / "output"):
                server.init_db()
                with server.connect() as connection:
                    connection.executemany(
                        """INSERT INTO findings(outage,unit_id,thimble_id,position,indication,percent,volts,location,channel,datapoint,analyst,imported_at)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                        [("H208", 2, 1, "B5", "WAR", 28, 1.2, "P1+10", "P1: 4-6", 80, "A", "now"),
                         ("H209", 2, 1, "B5", "WAR", 30, 1.4, "P1+10", "P1: 4-6", 80, "B", "now"),
                         ("H209", 2, 2, "C8", "WAR", 45, 2.1, "P4+20", "P3: 4-6", 90, "B", "now")],
                    )
                    selected_id = connection.execute("SELECT id FROM findings WHERE outage='H209' AND thimble_id=1").fetchone()[0]
                preview = server.build_inspection_preview("H209", 2, {}, [selected_id])
                self.assertEqual(preview["rows"], 1)
                result = server.export_inspection_docx("H209", 2, {}, [selected_id])
                with zipfile.ZipFile(result["file_path"]) as archive:
                    document_xml = archive.read("word/document.xml").decode("utf-8")
                self.assertIn("指套管涡流检验报告单", document_xml)
                comparison = server.export_comparison_docx("H208", "H209", 2)
                self.assertEqual(comparison["R"], 1)
                self.assertEqual(comparison["NI"], 1)
                self.assertTrue(Path(comparison["file_path"]).is_file())
                with zipfile.ZipFile(comparison["file_path"]) as archive:
                    comparison_xml = archive.read("word/document.xml").decode("utf-8")
                self.assertIn('<w:gridSpan w:val="3"/>', comparison_xml)
                self.assertIn("R</w:t>", comparison_xml)
                self.assertIn("结果比对栏内", comparison_xml)
                excel = server.export_comparison_excel("H208", "H209", 2)
                workbook = load_workbook(excel["file_path"])
                comparison_sheet = workbook.active
                self.assertIn("D2:F2", {str(item) for item in comparison_sheet.merged_cells.ranges})
                self.assertIn("G2:I2", {str(item) for item in comparison_sheet.merged_cells.ranges})
                self.assertTrue(str(comparison_sheet.cell(comparison_sheet.max_row, 1).value).startswith("备注："))

    def test_report_merges_only_matching_channel_and_core_position(self):
        rows = [
            {"thimble_id": 1, "position": "L11"},
            {"thimble_id": 1, "position": "L11"},
            {"thimble_id": 1, "position": "G14"},
            {"thimble_id": 2, "position": "L11"},
            {"thimble_id": 2, "position": "L11"},
            {"thimble_id": 3, "position": ""},
            {"thimble_id": 3, "position": ""},
        ]
        self.assertEqual(server.report_tube_rowspans(rows), [2, 0, 1, 2, 0, 1, 1])

        xml = server._w_table(
            ["序号", "通道编号", "堆芯位置"],
            [[1, 1, "L11"], [2, 1, "L11"], [3, 1, "G14"]],
            merge_group_columns=(1, 2), merge_columns=(1, 2),
        )
        self.assertEqual(xml.count('<w:vMerge w:val="restart"/>'), 2)
        self.assertEqual(xml.count("<w:vMerge/>") , 2)

    def test_export_query_is_not_truncated_at_ui_page_limit(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "export.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as connection:
                    connection.executemany(
                        "INSERT INTO findings(outage,unit_id,thimble_id,position,imported_at) VALUES(?,?,?,?,?)",
                        [("H209", 2, index % 50 + 1, "B5", "now") for index in range(205)],
                    )
                rows = server.query_all_findings({"unit": ["2"]})
            self.assertEqual(len(rows), 205)

    def test_state_payload_validation_rejects_invalid_values(self):
        with self.assertRaises(ValueError):
            server.validate_state_payload({"outage": "bad", "unit_id": 2, "thimble_id": 1, "state": "normal"})
        with self.assertRaises(ValueError):
            server.validate_state_payload({"outage": "H209", "unit_id": 2, "thimble_id": 51, "state": "normal"})
        with self.assertRaises(ValueError):
            server.validate_state_payload({"outage": "H209", "unit_id": 2, "thimble_id": 1, "state": "unknown"})

    def test_database_migration_backs_up_and_preserves_legacy_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "legacy.db"
            connection = sqlite3.connect(db)
            connection.execute("""CREATE TABLE findings (
                id INTEGER PRIMARY KEY, outage TEXT NOT NULL, unit_id INTEGER NOT NULL,
                thimble_id INTEGER NOT NULL, position TEXT, entry_no INTEGER, volts REAL,
                degrees REAL, indication TEXT, percent REAL, channel TEXT, location TEXT,
                datapoint INTEGER, liss_region_size INTEGER, analyst TEXT, analysis TEXT,
                filepath TEXT, filename TEXT, calgroup TEXT, report_path TEXT, imported_at TEXT NOT NULL)""")
            connection.execute("INSERT INTO findings(outage,unit_id,thimble_id,position,filename,calgroup,imported_at) VALUES('H209',2,3,'E11','DIR003C003I004.ECT','TH2I09CAL00001','now')")
            connection.commit(); connection.close()
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                with server.connect() as migrated:
                    row = migrated.execute("SELECT outage,unit_id,thimble_id,position,source_key FROM findings").fetchone()
                    columns = {item[1] for item in migrated.execute("PRAGMA table_info(findings)")}
            self.assertEqual(tuple(row), ("H209", 2, 3, "E11", None))
            self.assertIn("uid", columns)
            self.assertTrue(db.with_suffix(db.suffix + ".before-raw-fields.bak").is_file())

    def test_outage_ignores_calgroup_digits(self):
        path = Path(r"D:\data\H209\TH2I09CAL00005\Report-x.rpt")
        self.assertEqual(server.infer_outage(path), "H209")

    def test_parse_real_report(self):
        reports = [path for path in DATA_ROOT.rglob("Report*.rpt") if path.parent.name.upper().startswith("TH")]
        self.assertTrue(reports)
        findings = server.parse_report(reports[0])
        self.assertTrue(findings)
        self.assertTrue(all(1 <= x.thimble_id <= 50 for x in findings))
        self.assertTrue(all(x.position for x in findings))

    def test_import_is_idempotent(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "test.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                first = server.import_directory(str(H209_ROOT))
                second = server.import_directory(str(H209_ROOT))
            self.assertGreater(first["inserted"], 0)
            self.assertEqual(second["inserted"], 0)
            self.assertEqual(second["skipped"], first["parsed"])

    def test_overview_contains_compact_outage_history(self):
        with tempfile.TemporaryDirectory() as directory:
            with patch.object(server, "DB_PATH", Path(directory) / "history.db"):
                server.init_db()
                server.import_directory(str(H209_ROOT))
                history = server.overview()["history"]
                self.assertTrue(history)
                self.assertEqual({item["outage"] for item in history}, {"H209"})
                self.assertTrue(all(item["findings"] > 0 and item["tubes"] > 0 for item in history))

    def test_import_excludes_steam_generator_reports(self):
        with tempfile.TemporaryDirectory() as directory:
            db = Path(directory) / "test.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                server.import_directory(str(N208_ROOT))
                with server.connect() as connection:
                    maximum = connection.execute("SELECT MAX(thimble_id) FROM findings").fetchone()[0]
                self.assertLessEqual(maximum, 50)

    def test_ect_coverage_excludes_calibration_and_steam_generator(self):
        h209 = server.scan_thimble_directory(str(H209_ROOT))
        n208 = server.scan_thimble_directory(str(N208_ROOT))
        self.assertEqual(h209["unique_tubes"], list(range(1, 51)))
        self.assertEqual(n208["unique_tubes"], list(range(1, 51)))
        self.assertEqual(h209["missing_tubes"], [])
        self.assertEqual(n208["missing_tubes"], [])
        self.assertGreater(h209["calibration_count"], 0)
        self.assertGreater(n208["calibration_count"], 0)
        self.assertTrue(all(Path(row["path"]).parent.name.upper().startswith("TH") for row in n208["ect"]))

    def test_folder_exports_real_xlsx(self):
        with tempfile.TemporaryDirectory() as directory:
            with patch.object(server, "EXCEL_DIR", Path(directory)):
                result = server.export_directory_excel(str(H209_ROOT))
                target = Path(result["file_path"])
                self.assertTrue(target.is_file())
                book = load_workbook(target, read_only=True)
                required = ["缺陷明细表", "管子汇总表", "未进入RPT的真实ECT", "DIR999标定文件", "同名同哈希副本", "同名不同哈希文件", "RPT引用异常", "ECT编号校验异常", "SUM或机组号缺失", "堆芯映射失败"]
                self.assertEqual(book.sheetnames, required)
                sheet = book["缺陷明细表"]
                self.assertEqual(sheet.max_row - 1, result["rows"])
                headers = {cell.value: cell.column for cell in sheet[1]}
                self.assertEqual(sheet.cell(2, headers["电站名称"]).value, "红沿河")
                self.assertEqual(sheet.cell(2, headers["机组号"]).value, 2)
                self.assertGreater(book["管子汇总表"].max_row, 1)
                book.close()

    def test_imports_custom_standard_excel_by_header_name(self):
        with tempfile.TemporaryDirectory() as directory:
            source, db = Path(directory) / "custom.xlsx", Path(directory) / "test.db"
            book = Workbook()
            sheet = book.active
            sheet.title = "用户自定义数据"
            sheet.append(["备注", "数据组", "磨损位置", "分析人员", "通道编号", "机组号", "电站名称", "三字码", "磨损深度", "幅值", "相位"])
            sheet.append(["人工备注", "TH1I20CAL00001", "P1+40", "MQQ", 2, 1, "大亚湾", "WAR", 48, 5.07, 258])
            book.save(source)
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                result = server.import_excel_file(str(source))
                repeated = server.import_excel_file(str(source))
                self.assertEqual(result["parsed"], 1)
                self.assertEqual(repeated["inserted"], 0)
                self.assertEqual(repeated["skipped"], 1)
                with server.connect() as connection:
                    row = connection.execute("SELECT outage,unit_id,thimble_id,position,location FROM findings").fetchone()
                self.assertEqual(tuple(row), ("D120", 1, 2, "G14", "P1+40"))

    def test_imports_case_insensitive_english_template_without_column_shift(self):
        with tempfile.TemporaryDirectory() as directory:
            source, db = Path(directory) / "english.xlsx", Path(directory) / "test.db"
            book = Workbook(); sheet = book.active
            sheet.append(["Outage", "Unit Number", "Thimble ID", "Core Position", "Amplitude", "Phase", "Wear Percent", "Defect", "Measurement Channel", "Wear Location", "Analyst", "Filename", "Data Group", "Notes"])
            sheet.append(["H209", 2, 42, "M5", 5.07, 258, 22, "WAR", "P1: 4-6", "P1+10", "MQQ", "DIR042C042I001.ECT", "TH2I09CAL00001", "english import"])
            book.save(source); book.close()
            with patch.object(server, "DB_PATH", db):
                server.init_db(); result = server.import_excel_file(str(source))
                with server.connect() as connection:
                    row = connection.execute("SELECT outage,unit_id,thimble_id,position,percent,location,analyst,calgroup FROM findings").fetchone()
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(tuple(row), ("H209", 2, 42, "M5", 22, "P1+10", "MQQ", "TH2I09CAL00001"))

    def test_discovers_server_thimble_directories_without_fixed_depth(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "2--------红沿河项目" / "LHNP2" / "9.H209" / "2. 数据" / "2.3 指套管"
            (target / "TH2I09CAL00001").mkdir(parents=True)
            unrelated = root / "3--------宁德项目" / "NDNP1" / "N109" / "2.数据" / "2.1蒸发器"
            (unrelated / "SG1").mkdir(parents=True)
            result = server.discover_server_sources(str(root))
            self.assertEqual(len(result["items"]), 1)
            self.assertEqual(result["items"][0]["outage"], "H209")
            self.assertEqual(result["items"][0]["station"], "红沿河")
            self.assertEqual(result["items"][0]["unit_folder"], "LHNP2")
            self.assertEqual(Path(result["items"][0]["path"]), target)

    def test_infers_station_from_documented_project_and_unit_paths(self):
        samples = {
            r"DATA\1--------大亚湾项目\CNPS1\D110": "大亚湾",
            r"DATA\1--------大亚湾项目\LNPS2\L210": "岭澳",
            r"DATA\3--------宁德项目\NDNP4\N409": "宁德",
            r"DATA\4--------阳江项目\YJNP6\Y603": "阳江",
            r"DATA\5--------防城港项目\BLNP1\F107": "防城港",
            r"DATA\6--------台山项目\TSNP2\T205": "台山",
            r"DATA\7--------太平岭（惠州）项目\HENP1\P101": "太平岭（惠州）",
            r"DATA\8--------苍南项目\CNMP2\C201": "苍南",
        }
        for path, expected in samples.items():
            with self.subTest(path=path):
                self.assertEqual(server.infer_station(Path(path))[0], expected)

    def test_infers_outage_from_sum_when_parent_path_has_no_outage(self):
        report = DATA_ROOT / "指套管数据2" / "2.3 指套管" / "TH2I09CAL00001" / "Report-TH2I09CAL00001-KYY-SEC.rpt"
        self.assertEqual(server.infer_outage(report), "H209")

    def test_report_options_merge_byte_identical_copies(self):
        options = server.report_options(str(DATA_ROOT))
        h209 = [item for item in options if item.get("outage") == "H209" and item.get("group") == "TH2I09CAL00001"]
        self.assertEqual(len(h209), 1)
        self.assertGreaterEqual(len(h209[0]["duplicate_paths"]), 1)

    def test_report_option_distinguishes_entries_from_unique_tubes(self):
        options = server.report_options(str(DATA_ROOT / "指套管数据2"))
        y108 = next(item for item in options if item.get("outage") == "Y108" and item.get("group") == "TH1I08CAL00002" and "LYY" in item.get("name", ""))
        self.assertEqual(y108["records"], 41)
        self.assertEqual(y108["tube_count"], 24)
        self.assertEqual(y108["indication_count"], 32)

    def test_strict_group_processing_and_layered_counts(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); group = root / "TH2I09CAL00001"; group.mkdir()
            write_sum(group)
            write_ect(group, "DIR003C003I004.ECT", 3, 3, 4)
            write_ect(group, "DIR003C003I005.ECT", 3, 3, 5)
            write_ect(group, "DIR004C004I006.ECT", 5, 4, 6)
            write_ect(group, "DIR999C999I001.ECT", 999, 999, 1)
            write_report(group, [
                report_entry("DIR003C003I004.ECT", 3, 3, 4, "u1"),
                report_entry("DIR003C003I004.ECT", 3, 3, 4, "u2"),
                report_entry("DIR999C999I001.ECT", 999, 999, 1, "cal"),
                report_entry("DIR050C050I999.ECT", 50, 50, 999, "missing"),
            ])
            ignored = root / "整合一起看" / "TH2I09CAL00002"; ignored.mkdir(parents=True)
            write_sum(ignored); write_ect(ignored, "DIR001C001I001.ECT", 1, 1, 1)
            (root / "TH2I09CAL00001-copy").mkdir()
            result = server.process_directory(str(root))
            self.assertEqual(len(result["contexts"]), 1)
            self.assertEqual(len(result["report_rows"]), 4)
            self.assertEqual(len(result["findings"]), 2)
            tube = next(row for row in result["tube_summary"] if row["thimble_id"] == 3)
            self.assertEqual((tube["entry_count"], tube["raw_indication_count"], tube["position_group_count"]), (2, 2, 1))
            self.assertEqual(tube["position"], "E11")
            self.assertEqual(len(result["calibration_ect"]), 1)
            self.assertEqual(len(result["ect_mismatches"]), 1)
            self.assertEqual(len(result["report_reference_errors"]), 2)
            self.assertEqual(len(result["unreported_ect"]), 2)
            db = root / "test.db"
            with patch.object(server, "DB_PATH", db):
                server.init_db()
                imported = server.import_directory(str(root))
                self.assertEqual(imported["inserted"], 2)
                with server.connect() as connection:
                    rows = connection.execute("SELECT degrees,uid,measurement_type,distance,extent,length,width,comment FROM findings ORDER BY uid").fetchall()
                self.assertEqual([row["uid"] for row in rows], ["u1", "u2"])
                self.assertTrue(all(row["degrees"] == 0 for row in rows))
                self.assertTrue(all(tuple(row[key] for key in ("measurement_type", "distance", "extent", "length", "width", "comment")) == ("Point", "10", "2", "3", "4", "note") for row in rows))

    def test_missing_sum_never_guesses_unit_or_position(self):
        with tempfile.TemporaryDirectory() as directory:
            group = Path(directory) / "TH2I09CAL00001"; group.mkdir()
            write_ect(group, "DIR003C003I004.ECT", 3, 3, 4)
            write_report(group, [report_entry("DIR003C003I004.ECT", 3, 3, 4, "u1")])
            result = server.process_directory(directory)
            self.assertEqual(result["findings"][0].unit_id, 0)
            self.assertEqual(result["findings"][0].position, "")
            self.assertTrue(result["sum_warnings"])
            self.assertTrue(result["mapping_failures"])

    def test_physical_duplicate_hash_classification(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            for branch, marker in (("a", "same"), ("b", "same"), ("c", "different")):
                group = root / branch / "TH2I09CAL00001"; group.mkdir(parents=True)
                write_sum(group); write_ect(group, "DIR003C003I004.ECT", 3, 3, 4, marker)
            result = server.process_directory(directory)
            self.assertEqual(len(result["duplicate_same_hash"]), 1)
            self.assertEqual(len(result["same_name_different_hash"]), 3)


if __name__ == "__main__": unittest.main()
