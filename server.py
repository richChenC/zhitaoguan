from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import logging
import mimetypes
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

VENDOR_DIR = Path(__file__).resolve().parent / ".vendor"
if VENDOR_DIR.is_dir():
    sys.path.insert(0, str(VENDOR_DIR))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
DATA_DIR = Path(os.environ.get("THIMBLE_DATA_DIR", str(ROOT / "data"))).expanduser().resolve()
DB_PATH = Path(os.environ.get("THIMBLE_DB_PATH", str(DATA_DIR / "thimble.db"))).expanduser().resolve()
EXCEL_DIR = Path(os.environ.get("THIMBLE_OUTPUT_DIR", str(ROOT / "output" / "excel"))).expanduser().resolve()
REPORT_DIR = Path(os.environ.get("THIMBLE_REPORT_DIR", str(ROOT / "output" / "reports"))).expanduser().resolve()
LOG_PATH = Path(os.environ.get("THIMBLE_LOG_PATH", str(DATA_DIR / "thimble.log"))).expanduser().resolve()
SERVICE_VERSION = "2026.08.06"
try:
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(filename=LOG_PATH, level=logging.INFO, encoding="utf-8", format="%(asctime)s %(levelname)s %(message)s")
except OSError:
    logging.basicConfig(stream=sys.stderr, level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
LOGGER = logging.getLogger("thimble")

ODD_POSITIONS = "L11 G14 N7 H13 J12 R8 N12 N10 L14 J15 H11 F13 J7 L5 M5 L8 N8 L6 J10 L9 F9 C12 G7 L4 J5 M3 G9 E11 F11 D12 F6 B10 D7 E5 H3 J3 H6 H4 F8 D10 B7 B5 D3 D5 F2 H1 B8 F4 C8 A9".split()
EVEN_POSITIONS = "B5 C8 E11 D10 D12 C12 B10 B7 A9 B8 D5 D3 F6 H13 F9 G14 F13 E5 F11 D7 G7 F2 H6 H11 J10 J12 J15 G9 F8 F4 H3 H1 J3 N12 L9 L11 L14 J5 J7 H4 M3 M5 R8 N7 N8 N10 L5 L8 L6 L4".split()

DATA_GROUP_NAME = re.compile(r"^TH\d+I\d+CAL\d+$", re.I)
IGNORED_SCAN_DIRECTORIES = {
    "整合一起看", "临时输出", "临时文件", "输出", "output", "excel", "tmp", "temp",
    "人工汇总", "人工汇总表", "汇总表", "辅助目录",
}
SITE_NAMES = {
    "LHNP": "红沿河", "CNPS": "大亚湾", "LNPS": "岭澳", "NDNP": "宁德",
    "YJNP": "阳江", "BLNP": "防城港", "TSNP": "台山", "HENP": "太平岭",
    "CNMP": "苍南", "LFNP": "陆丰",
}

STANDARD_EXCEL_FIELDS = [
    "电站名称", "大修号", "机组号", "安全级别", "设备名称", "被检部位", "通道编号",
    "在堆芯内的位置", "幅值", "相位", "磨损深度", "三字码", "测量通道",
    "磨损位置", "探头类型", "分析人员", "数据", "数据组", "备注", "堵管移位信息"
]
STATION_CODES = {
    "大亚湾": "D", "岭澳": "L", "宁德": "N", "红沿河": "H", "阳江": "Y",
    "防城港": "F", "台山": "T", "太平岭（惠州）": "P", "苍南": "C", "陆丰": "L",
}
STATION_PROJECT_HINTS = (
    ("大亚湾", ("CNPS",)),
    ("岭澳", ("LNPS",)),
    ("红沿河", ("红沿河项目", "LHNP")),
    ("宁德", ("宁德项目", "NDNP")),
    ("阳江", ("阳江项目", "YJNP")),
    ("防城港", ("防城港项目", "BLNP")),
    ("台山", ("台山项目", "TSNP")),
    ("太平岭（惠州）", ("太平岭", "惠州", "HENP")),
    ("苍南", ("苍南项目", "CNMP")),
    ("陆丰", ("陆丰项目", "LFNP")),
    ("外部市场", ("外部市场",)),
)


@dataclass
class Finding:
    outage: str
    unit_id: int
    thimble_id: int
    position: str
    entry_no: int | None
    volts: float | None
    degrees: float | None
    indication: str
    percent: float | None
    channel: str
    location: str
    datapoint: int | None
    liss_region_size: int | None
    analyst: str
    analysis: str
    filepath: str
    filename: str
    calgroup: str
    report_path: str
    site_code: str = ""
    site_owner: str = ""
    component: str = ""
    tester: str = ""
    probe: str = ""
    column_no: int | None = None
    channel_id: str = ""
    measurement_type: str = ""
    uid: str = ""
    distance: str = ""
    extent: str = ""
    length: str = ""
    width: str = ""
    comment: str = ""
    source_key: str = ""
    business_key: str = ""
    source_indication: str = ""
    indication_resolution: str = "source"


BUSINESS_KEY_FIELDS = (
    "outage", "unit_id", "thimble_id", "position", "entry_no", "volts", "degrees",
    "indication", "percent", "channel", "location", "datapoint", "liss_region_size",
    "analyst", "analysis", "filename", "calgroup", "site_code", "site_owner", "component",
    "tester", "probe", "column_no", "channel_id", "measurement_type", "uid", "distance",
    "extent", "length", "width", "comment",
)


def finding_business_key(values) -> str:
    """Identity of one report record, excluding import time and machine paths."""
    def read(name):
        value_ = values.get(name) if isinstance(values, dict) else getattr(values, name, None)
        return "" if value_ is None else str(value_).strip().casefold()
    payload = "\x1f".join(read(name) for name in BUSINESS_KEY_FIELDS)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def audit_duplicate(db: sqlite3.Connection, values: dict, action: str = "skipped") -> None:
    business_key = values.get("business_key") or finding_business_key(values)
    kept = db.execute("SELECT id FROM findings WHERE business_key=? ORDER BY id LIMIT 1", (business_key,)).fetchone()
    db.execute("""INSERT INTO duplicate_records(
        detected_at,kept_finding_id,duplicate_source_key,business_key,reason,source_path,payload_json,action
    ) VALUES(?,?,?,?,?,?,?,?)""", (
        datetime.now().isoformat(timespec="seconds"), kept["id"] if kept else None,
        values.get("source_key", ""), business_key, "完全相同的检测记录",
        values.get("report_path") or values.get("filepath") or "",
        json.dumps(values, ensure_ascii=False, default=str), action,
    ))


def duplicate_summary(limit: int = 100) -> dict:
    with connect() as db:
        total = db.execute("SELECT COUNT(*) FROM findings").fetchone()[0]
        duplicate_count = db.execute("SELECT COUNT(*) FROM duplicate_records").fetchone()[0]
        rows = db.execute("""SELECT id,detected_at,kept_finding_id,reason,source_path,action
            FROM duplicate_records ORDER BY id DESC LIMIT ?""", (max(1, min(limit, 500)),)).fetchall()
    return {"findings": total, "duplicates": duplicate_count, "items": [dict(row) for row in rows]}


def deduplicate_database() -> dict:
    init_db()
    return duplicate_summary()


@dataclass(frozen=True)
class DataGroupContext:
    path: Path
    calgroup: str
    site_owner: str = ""
    site_code: str = ""
    unit_id: int = 0
    component: str = ""
    outage: str = ""
    probe_model: str = ""
    tester: str = ""
    sum_path: str = ""
    warning: str = ""


def value(node: ET.Element, name: str) -> str:
    child = node.find(name)
    return (child.text or "").strip() if child is not None else ""


def number(text: str, cast):
    try:
        return cast(text)
    except (TypeError, ValueError):
        return None


def infer_outage(path: Path) -> str:
    for part in reversed(path.parts):
        match = re.search(r"(?<![A-Z0-9])([A-Z]\d{3})(?![A-Z0-9])", part.upper())
        if match:
            return match.group(1)
    directory = path if path.is_dir() else path.parent
    try:
        summaries = [candidate for candidate in directory.iterdir() if candidate.is_file() and candidate.suffix.lower() == ".sum"]
    except (OSError, PermissionError):
        summaries = []
    for summary in summaries:
        try:
            site_node = ET.parse(summary).getroot().find("Site")
            outage = value(site_node, "Outage").upper() if site_node is not None else ""
            if re.fullmatch(r"[A-Z]\d{3}", outage):
                return outage
        except (ET.ParseError, OSError):
            continue
    return "UNKNOWN"


def infer_unit(outage: str, report_id: str) -> int:
    match = re.search(r"[A-Z](\d)", outage)
    return int(match.group(1)) if match else (number(report_id, int) or 0)


def infer_station(path: Path, outage: str = "") -> tuple[str, str]:
    context = "|".join(path.parts).upper()
    for station, hints in STATION_PROJECT_HINTS:
        if any(hint.upper() in context for hint in hints):
            return station, STATION_CODES.get(station, outage[:1])
    code = outage[:1].upper()
    # 大亚湾项目是项目级目录，内部仍需按 D/L 区分大亚湾与岭澳电站。
    if "大亚湾项目" in context:
        station = {"D": "大亚湾", "L": "岭澳"}.get(code, "大亚湾")
        return station, code or STATION_CODES[station]
    station = next((name for name, station_code in STATION_CODES.items() if station_code == code), code or "未知基地")
    return station, code


def position_for(unit_id: int, thimble_id: int) -> str:
    positions = ODD_POSITIONS if unit_id % 2 else EVEN_POSITIONS
    return positions[thimble_id - 1] if 1 <= thimble_id <= len(positions) else ""


def split_location(location: str) -> tuple[str, float | None]:
    match = re.match(r"^\s*(P\d+)\s*(?:\+\s*([-+]?\d+(?:\.\d+)?))?\s*$", location or "", re.I)
    return (match.group(1).upper(), number(match.group(2), float)) if match else ((location or "").strip(), None)


def format_location(location: str) -> str:
    zone, offset = split_location(location)
    if not zone:
        return "-"
    if not re.fullmatch(r"P[1-6]", zone, re.I):
        return zone
    value = 0 if offset is None else offset
    rendered = str(int(value)) if float(value).is_integer() else f"{value:g}"
    return f"{zone.upper()} + {rendered} mm"


def normalize_header(value_) -> str:
    return re.sub(r"[\s()（）%％/\\_-]+", "", str(value_ or "")).lower()


def outage_from_standard_row(row: dict, source: Path) -> str:
    explicit = str(row.get("大修号") or row.get("大修") or "").strip().upper()
    if re.fullmatch(r"[A-Z]\d{3}", explicit):
        return explicit
    group = str(row.get("数据组") or "").strip().upper()
    match = re.search(r"TH(\d)I(\d{2})", group)
    station = str(row.get("电站名称") or "").strip()
    code = next((value_ for name, value_ in STATION_CODES.items() if name in station), "")
    if match and code:
        return f"{code}{match.group(1)}{match.group(2)}"
    inferred = infer_outage(source)
    return inferred if inferred != "UNKNOWN" else ""


ECT_NAME = re.compile(r"^DIR(\d{3})C(\d{3})I(\d{3})\.ECT$", re.I)


def discover_data_groups(directory: str | Path, ignored_directories: set[str] | None = None) -> list[Path]:
    base = Path(directory).expanduser().resolve()
    if not base.is_dir():
        raise ValueError("所选路径不是有效文件夹")
    ignored = {name.casefold() for name in (ignored_directories or set()) | IGNORED_SCAN_DIRECTORIES}
    groups: list[Path] = []
    for current, directories, _ in os.walk(base):
        directories[:] = [name for name in directories if name.casefold() not in ignored]
        path = Path(current)
        if DATA_GROUP_NAME.fullmatch(path.name):
            groups.append(path)
            directories[:] = []
    return sorted(groups, key=lambda path: str(path).casefold())


def parse_sum(group_dir: Path) -> DataGroupContext:
    summaries = sorted(path for path in group_dir.iterdir() if path.is_file() and path.suffix.casefold() == ".sum")
    if not summaries:
        return DataGroupContext(group_dir, group_dir.name, warning="缺少 SUM 文件，机组号和堆芯位置留空")
    warnings = []
    if len(summaries) > 1:
        warnings.append(f"发现 {len(summaries)} 个 SUM，使用 {summaries[0].name}")
    try:
        root = ET.parse(summaries[0]).getroot()
        site = root.find("Site"); site = site if site is not None else root
        probe = root.find("Probe"); probe = probe if probe is not None else root
        equipment = root.find("Equipment"); equipment = equipment if equipment is not None else root
        unit_id = number(value(site, "Unit"), int) or 0
        if unit_id <= 0:
            warnings.append("SUM Unit 缺失或无法解析，堆芯位置留空")
        component = value(site, "Component")
        if component and component.upper() != "TH":
            warnings.append(f"SUM Component={component}，不是 TH")
        return DataGroupContext(
            path=group_dir, calgroup=group_dir.name, site_owner=value(site, "Owner"),
            site_code=value(site, "SiteCode"), unit_id=unit_id, component=component,
            outage=value(site, "Outage").upper(), probe_model=value(probe, "Model"),
            tester=value(equipment, "Tester"), sum_path=str(summaries[0]), warning="；".join(warnings),
        )
    except (ET.ParseError, OSError) as exc:
        return DataGroupContext(group_dir, group_dir.name, sum_path=str(summaries[0]), warning=f"SUM 解析失败：{exc}")


def parse_ect_header(path: Path) -> dict:
    match = ECT_NAME.match(path.name)
    if not match:
        raise ValueError("ECT文件名不符合 DIRrrrCcccIeee.ECT")
    name_row, name_col, name_entry = map(int, match.groups())
    raw = path.read_bytes()[:131072].decode("utf-8", errors="ignore")
    xml_match = re.search(r"<HeaderTube\b.*?</HeaderTube>", raw, re.S)
    if not xml_match:
        raise ValueError("ECT中未找到HeaderTube XML")
    header = ET.fromstring(xml_match.group(0))
    row, col, entry = number(value(header, "Row"), int), number(value(header, "Col"), int), number(value(header, "Entry"), int)
    return {
        "path": str(path), "filename": path.name, "calgroup": path.parent.name,
        "name_row": name_row, "name_col": name_col, "name_entry": name_entry,
        "row": row, "col": col, "entry": entry, "datetime": value(header, "DateTime"),
        "calibration": name_row == 999 or row == 999,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
    }


def scan_thimble_directory(directory: str) -> dict:
    base = Path(directory).expanduser().resolve()
    if not base.is_dir():
        raise ValueError("所选路径不是有效文件夹")
    ect_rows, errors = [], []
    paths = [path for group in discover_data_groups(base) for path in group.iterdir() if path.is_file() and path.suffix.casefold() == ".ect"]
    for path in sorted(paths):
        try:
            item = parse_ect_header(path)
            status = []
            if (item["name_row"], item["name_col"], item["name_entry"]) != (item["row"], item["col"], item["entry"]):
                status.append("文件名与HeaderTube不一致")
            if not item["calibration"] and not 1 <= (item["row"] or 0) <= 50:
                status.append("管号超出1-50")
            item["validation"] = "通过" if not status else "；".join(status)
            ect_rows.append(item)
            if status:
                errors.append({"file": str(path), "type": "ECT校验", "message": item["validation"]})
        except Exception as exc:
            errors.append({"file": str(path), "type": "ECT解析", "message": str(exc)})
    normal = [item for item in ect_rows if not item["calibration"] and item["validation"] == "通过"]
    unique_tubes = sorted({item["row"] for item in normal})
    return {
        "ect": ect_rows, "errors": errors, "unique_tubes": unique_tubes,
        "missing_tubes": sorted(set(range(1, 51)) - set(unique_tubes)),
        "calibration_count": sum(item["calibration"] for item in ect_rows),
    }


def parse_report(path: Path, context: DataGroupContext | None = None) -> list[Finding]:
    root = ET.parse(path).getroot()
    context = context or parse_sum(path.parent)
    findings: list[Finding] = []
    for item in root.findall("ReportEntry"):
        thimble_id = number(value(item, "Row"), int) or 0
        uid = value(item, "Uid")
        filename = value(item, "Filename")
        calgroup = value(item, "Calgroup") or context.calgroup
        source_identity = "|".join((context.site_code, str(context.unit_id), context.outage, calgroup, filename, uid or value(item, "Datapoint"), value(item, "Channel")))
        findings.append(Finding(
            outage=context.outage, unit_id=context.unit_id, thimble_id=thimble_id,
            position=position_for(context.unit_id, thimble_id) if context.unit_id else "",
            entry_no=number(value(item, "Entry"), int),
            volts=number(value(item, "Volts"), float),
            degrees=number(value(item, "Degrees"), float),
            indication=value(item, "Indication"), percent=number(value(item, "Percent"), float),
            channel=value(item, "Channel"), location=value(item, "Location"),
            datapoint=number(value(item, "Datapoint"), int),
            liss_region_size=number(value(item, "LissRegionSize"), int),
            analyst=value(item, "Analyst"), analysis=value(item, "Analysis"),
            filepath=value(item, "Filepath"), filename=filename,
            calgroup=calgroup, report_path=str(path), site_code=context.site_code,
            site_owner=context.site_owner, component=context.component, tester=context.tester,
            probe=value(item, "Probe") or context.probe_model,
            column_no=number(value(item, "Column"), int), channel_id=value(item, "ChannelID"),
            measurement_type=value(item, "MeasurementType"), uid=uid,
            distance=value(item, "Distance"), extent=value(item, "Extent"),
            length=value(item, "Length"), width=value(item, "Width"), comment=value(item, "Comment"),
            source_key=hashlib.sha256(source_identity.encode("utf-8")).hexdigest(),
        ))
    return findings


class ClosingConnection(sqlite3.Connection):
    def __exit__(self, exc_type, exc_value, traceback):
        try:
            return super().__exit__(exc_type, exc_value, traceback)
        finally:
            self.close()


def connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    db = sqlite3.connect(DB_PATH, factory=ClosingConnection)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys=ON")
    db.execute("PRAGMA busy_timeout=5000")
    db.execute("PRAGMA synchronous=NORMAL")
    return db


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.is_file():
        probe = sqlite3.connect(DB_PATH)
        try:
            columns = {row[1] for row in probe.execute("PRAGMA table_info(findings)")}
        finally:
            probe.close()
        if columns and "source_key" not in columns:
            backup = DB_PATH.with_suffix(DB_PATH.suffix + ".before-raw-fields.bak")
            if not backup.exists():
                shutil.copy2(DB_PATH, backup)
        if columns and "business_key" not in columns:
            backup = DB_PATH.with_suffix(DB_PATH.suffix + ".before-dedup.bak")
            if not backup.exists():
                shutil.copy2(DB_PATH, backup)
    with connect() as db:
        # WAL keeps the desktop UI responsive while an import or export is running.
        db.execute("PRAGMA journal_mode=WAL")
        db.executescript("""
        CREATE TABLE IF NOT EXISTS findings (
          id INTEGER PRIMARY KEY, outage TEXT NOT NULL, unit_id INTEGER NOT NULL,
          thimble_id INTEGER NOT NULL, position TEXT, entry_no INTEGER, volts REAL,
          degrees REAL, indication TEXT, percent REAL, channel TEXT, location TEXT,
          datapoint INTEGER, liss_region_size INTEGER, analyst TEXT, analysis TEXT,
          filepath TEXT, filename TEXT, calgroup TEXT, report_path TEXT,
          site_code TEXT DEFAULT '', site_owner TEXT DEFAULT '', component TEXT DEFAULT '',
          tester TEXT DEFAULT '', probe TEXT DEFAULT '', column_no INTEGER, channel_id TEXT DEFAULT '',
          measurement_type TEXT DEFAULT '', uid TEXT DEFAULT '', distance TEXT DEFAULT '', extent TEXT DEFAULT '',
          length TEXT DEFAULT '', width TEXT DEFAULT '', comment TEXT DEFAULT '', source_key TEXT,
          business_key TEXT, source_indication TEXT DEFAULT '', indication_resolution TEXT DEFAULT 'source',
          imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS tube_states (
          outage TEXT NOT NULL, unit_id INTEGER NOT NULL, thimble_id INTEGER NOT NULL,
          state TEXT NOT NULL DEFAULT 'normal', offset_mm REAL NOT NULL DEFAULT 0,
          note TEXT NOT NULL DEFAULT '', PRIMARY KEY(outage, unit_id, thimble_id)
        );
        CREATE TABLE IF NOT EXISTS duplicate_records (
          id INTEGER PRIMARY KEY, detected_at TEXT NOT NULL, kept_finding_id INTEGER,
          duplicate_source_key TEXT, business_key TEXT NOT NULL, reason TEXT NOT NULL,
          source_path TEXT DEFAULT '', payload_json TEXT NOT NULL, action TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS indication_resolution_audit (
          id INTEGER PRIMARY KEY, finding_id INTEGER NOT NULL, changed_at TEXT NOT NULL,
          source_indication TEXT DEFAULT '', resolved_indication TEXT NOT NULL,
          action TEXT NOT NULL, source_path TEXT DEFAULT ''
        );
        """)
        columns = {row[1] for row in db.execute("PRAGMA table_info(findings)")}
        if "source_key" not in columns:
            db.executescript("""
            ALTER TABLE findings RENAME TO findings_legacy;
            CREATE TABLE findings (
              id INTEGER PRIMARY KEY, outage TEXT NOT NULL, unit_id INTEGER NOT NULL,
              thimble_id INTEGER NOT NULL, position TEXT, entry_no INTEGER, volts REAL,
              degrees REAL, indication TEXT, percent REAL, channel TEXT, location TEXT,
              datapoint INTEGER, liss_region_size INTEGER, analyst TEXT, analysis TEXT,
              filepath TEXT, filename TEXT, calgroup TEXT, report_path TEXT,
              site_code TEXT DEFAULT '', site_owner TEXT DEFAULT '', component TEXT DEFAULT '',
              tester TEXT DEFAULT '', probe TEXT DEFAULT '', column_no INTEGER, channel_id TEXT DEFAULT '',
              measurement_type TEXT DEFAULT '', uid TEXT DEFAULT '', distance TEXT DEFAULT '', extent TEXT DEFAULT '',
              length TEXT DEFAULT '', width TEXT DEFAULT '', comment TEXT DEFAULT '', source_key TEXT,
              business_key TEXT, source_indication TEXT DEFAULT '', indication_resolution TEXT DEFAULT 'source',
              imported_at TEXT NOT NULL
            );
            INSERT INTO findings (
              id,outage,unit_id,thimble_id,position,entry_no,volts,degrees,indication,percent,
              channel,location,datapoint,liss_region_size,analyst,analysis,filepath,filename,
              calgroup,report_path,imported_at
            ) SELECT id,outage,unit_id,thimble_id,position,entry_no,volts,degrees,indication,percent,
              channel,location,datapoint,liss_region_size,analyst,analysis,filepath,filename,
              calgroup,report_path,imported_at FROM findings_legacy;
            DROP TABLE findings_legacy;
            """)
        columns = {row[1] for row in db.execute("PRAGMA table_info(findings)")}
        if "business_key" not in columns:
            db.execute("ALTER TABLE findings ADD COLUMN business_key TEXT")
        columns = {row[1] for row in db.execute("PRAGMA table_info(findings)")}
        if "source_indication" not in columns:
            db.execute("ALTER TABLE findings ADD COLUMN source_indication TEXT DEFAULT ''")
            db.execute("UPDATE findings SET source_indication=COALESCE(indication,'')")
        if "indication_resolution" not in columns:
            db.execute("ALTER TABLE findings ADD COLUMN indication_resolution TEXT DEFAULT 'source'")
            db.execute("""UPDATE findings SET indication_resolution='unreviewed'
                WHERE TRIM(COALESCE(indication,''))='' AND
                (TRIM(COALESCE(location,''))<>'' OR UPPER(TRIM(COALESCE(measurement_type,''))) NOT IN ('','NONE'))""")
        db.execute("DROP INDEX IF EXISTS idx_findings_business_key")
        rows = db.execute("SELECT * FROM findings WHERE business_key IS NULL OR business_key='' ORDER BY id").fetchall()
        for row in rows:
            db.execute("UPDATE findings SET business_key=? WHERE id=?", (finding_business_key(dict(row)), row["id"]))
        duplicates = db.execute("""SELECT business_key, MIN(id) kept_id, GROUP_CONCAT(id) ids
            FROM findings WHERE business_key IS NOT NULL AND business_key<>''
            GROUP BY business_key HAVING COUNT(*)>1""").fetchall()
        for duplicate in duplicates:
            for duplicate_id in (int(value_) for value_ in duplicate["ids"].split(",")):
                if duplicate_id == duplicate["kept_id"]: continue
                row = db.execute("SELECT * FROM findings WHERE id=?", (duplicate_id,)).fetchone()
                db.execute("""INSERT INTO duplicate_records(
                    detected_at,kept_finding_id,duplicate_source_key,business_key,reason,source_path,payload_json,action
                ) VALUES(?,?,?,?,?,?,?,?)""", (datetime.now().isoformat(timespec="seconds"), duplicate["kept_id"],
                    row["source_key"], duplicate["business_key"], "完全相同的检测记录", row["report_path"] or row["filepath"] or "",
                    json.dumps(dict(row), ensure_ascii=False, default=str), "removed"))
                db.execute("DELETE FROM findings WHERE id=?", (duplicate_id,))
        db.execute("DELETE FROM findings WHERE thimble_id NOT BETWEEN 1 AND 50")
        # Query paths are consistently scoped by outage, unit and tube; keep these
        # indexes local to SQLite so imports remain portable and fast.
        db.executescript("""
        CREATE INDEX IF NOT EXISTS idx_findings_scope ON findings(outage, unit_id, thimble_id);
        CREATE INDEX IF NOT EXISTS idx_findings_unit_tube ON findings(unit_id, thimble_id, outage);
        CREATE INDEX IF NOT EXISTS idx_findings_location ON findings(unit_id, thimble_id, location);
        DROP INDEX IF EXISTS idx_findings_source_key;
        CREATE UNIQUE INDEX idx_findings_source_key ON findings(source_key) WHERE source_key IS NOT NULL AND source_key<>'';
        CREATE UNIQUE INDEX IF NOT EXISTS idx_findings_business_key ON findings(business_key) WHERE business_key IS NOT NULL AND business_key<>'';
        CREATE INDEX IF NOT EXISTS idx_duplicate_records_detected ON duplicate_records(detected_at DESC);
        CREATE INDEX IF NOT EXISTS idx_indication_resolution_audit_finding ON indication_resolution_audit(finding_id,changed_at DESC);
        CREATE INDEX IF NOT EXISTS idx_findings_site_scope ON findings(site_code, unit_id, outage, thimble_id);
        CREATE INDEX IF NOT EXISTS idx_tube_states_scope ON tube_states(outage, unit_id, thimble_id);
        """)


def is_indication_record(finding: Finding) -> bool:
    measurement = finding.measurement_type.strip().casefold()
    return bool(finding.indication.strip() or finding.location.strip() or (measurement and measurement != "none"))


def blank_indication_summary(findings: list[Finding], limit: int = 100) -> dict:
    rows = [finding for finding in findings if not finding.indication.strip() and is_indication_record(finding)]
    samples = [{
        "outage": row.outage, "unit_id": row.unit_id, "thimble_id": row.thimble_id,
        "position": row.position, "percent": row.percent, "location": row.location,
        "channel": row.channel, "filename": row.filename, "calgroup": row.calgroup,
        "analyst": row.analyst, "source": row.report_path,
    } for row in rows[:limit]]
    return {"count": len(rows), "tubes": len({(row.outage, row.unit_id, row.thimble_id) for row in rows}),
            "sources": len({row.report_path for row in rows}), "items": samples}


def indication_columns(finding: Finding, policy: str) -> dict:
    columns = asdict(finding)
    source = finding.indication.strip()
    columns["source_indication"] = source
    if not source and is_indication_record(finding):
        columns["indication_resolution"] = "bulk_war" if policy == "fill_war" else "review"
        if policy == "fill_war":
            columns["indication"] = "WAR"
    else:
        columns["indication_resolution"] = "source"
    return columns


def audit_indication_resolution(db: sqlite3.Connection, finding_id: int, columns: dict) -> None:
    if columns.get("indication_resolution") != "bulk_war":
        return
    exists = db.execute("SELECT 1 FROM indication_resolution_audit WHERE finding_id=? AND action='bulk_war'", (finding_id,)).fetchone()
    if not exists:
        db.execute("""INSERT INTO indication_resolution_audit(
            finding_id,changed_at,source_indication,resolved_indication,action,source_path
        ) VALUES(?,?,?,?,?,?)""", (finding_id, datetime.now().isoformat(timespec="seconds"),
            columns.get("source_indication", ""), "WAR", "bulk_war",
            columns.get("report_path") or columns.get("filepath") or ""))


def display_indication(row) -> str:
    def get(name, default=None):
        try:
            return row[name]
        except (KeyError, IndexError, TypeError):
            return getattr(row, name, default)
    indication = str(get("indication", "") or "").strip()
    if indication.upper() in {"NDD", "NONE", "NO DEFECT"}:
        return "NDD"
    measurement = str(get("measurement_type", "") or "").strip().upper()
    has_measurement = bool(indication or str(get("location", "") or "").strip() or (measurement and measurement != "NONE"))
    return indication or ("未标注" if has_measurement else "NDD")


def process_directory(directory: str, selected_reports: list[str] | None = None) -> dict:
    base = Path(directory).expanduser().resolve()
    groups = discover_data_groups(base)
    selected = None if selected_reports is None else {str(Path(path).resolve()).casefold() for path in selected_reports}
    contexts = [parse_sum(group) for group in groups]
    result = {
        "contexts": contexts, "findings": [], "report_rows": [], "tube_summary": [],
        "unreported_ect": [], "calibration_ect": [], "duplicate_same_hash": [],
        "same_name_different_hash": [], "report_reference_errors": [], "ect_mismatches": [],
        "sum_warnings": [], "mapping_failures": [], "parse_errors": [], "ect": [], "reports": [],
    }
    ect_by_group: dict[tuple[str, str], list[dict]] = {}
    ect_by_name: dict[str, list[dict]] = {}
    physical_versions: dict[tuple, list[dict]] = {}
    report_fingerprints: set[tuple] = set()
    for context in contexts:
        if context.warning:
            result["sum_warnings"].append({"group": context.calgroup, "path": str(context.path), "sum": context.sum_path, "message": context.warning})
        for ect_path in sorted(path for path in context.path.iterdir() if path.is_file() and path.suffix.casefold() == ".ect"):
            try:
                ect = parse_ect_header(ect_path)
                ect["context"] = context
                ect["logical_key"] = (context.site_code, context.unit_id, context.component.upper(), context.outage, context.calgroup.upper(), ect["filename"].upper())
                ect["header_matches_name"] = (ect["name_row"], ect["name_col"], ect["name_entry"]) == (ect["row"], ect["col"], ect["entry"])
                result["ect"].append(ect)
                ect_by_group.setdefault((str(context.path).casefold(), ect["filename"].upper()), []).append(ect)
                ect_by_name.setdefault(ect["filename"].upper(), []).append(ect)
                physical_versions.setdefault(ect["logical_key"], []).append(ect)
                if ect["calibration"]:
                    result["calibration_ect"].append(ect)
                if not ect["header_matches_name"]:
                    result["ect_mismatches"].append(ect)
            except Exception as exc:
                result["parse_errors"].append({"type": "ECT 解析", "file": str(ect_path), "message": str(exc)})
        reports = sorted(path for path in context.path.iterdir() if path.is_file() and path.suffix.casefold() == ".rpt")
        for report in reports:
            if selected is not None and str(report.resolve()).casefold() not in selected:
                continue
            try:
                fingerprint = (context.site_code.upper(), context.unit_id, context.component.upper(), context.outage, context.calgroup.upper(), hashlib.sha256(report.read_bytes()).hexdigest())
            except OSError:
                fingerprint = (context.site_code.upper(), context.unit_id, context.component.upper(), context.outage, context.calgroup.upper(), str(report.resolve()).casefold())
            if fingerprint in report_fingerprints:
                continue
            report_fingerprints.add(fingerprint)
            result["reports"].append(report)
            try:
                rows = parse_report(report, context)
                result["report_rows"].extend(rows)
            except Exception as exc:
                result["parse_errors"].append({"type": "RPT 解析", "file": str(report), "message": str(exc)})
    for logical_key, copies in physical_versions.items():
        hashes: dict[str, list[dict]] = {}
        for ect in copies:
            hashes.setdefault(ect["sha256"], []).append(ect)
        for same_hash in hashes.values():
            if len(same_hash) > 1:
                result["duplicate_same_hash"].extend(same_hash[1:])
        if len(hashes) > 1:
            result["same_name_different_hash"].extend(copies)
    referenced_paths: set[str] = set()
    for finding in result["report_rows"]:
        group_key = (str(Path(finding.report_path).parent.resolve()).casefold(), finding.filename.upper())
        candidates = ect_by_group.get(group_key, [])
        error = ""
        if finding.calgroup.upper() != Path(finding.report_path).parent.name.upper():
            error = "RPT Calgroup 与当前数据组不一致"
        elif not candidates:
            elsewhere = ect_by_name.get(finding.filename.upper(), [])
            error = "RPT 跨组引用 ECT" if elsewhere else "RPT 引用的 ECT 不存在"
        else:
            ect = candidates[0]
            referenced_paths.update(item["path"].casefold() for item in candidates)
            if ect["calibration"]:
                error = "RPT 引用了 DIR999 标定文件"
            elif (finding.thimble_id, finding.column_no, finding.entry_no) != (ect["row"], ect["col"], ect["entry"]):
                error = "RPT Row/Column/Entry 与 ECT 不一致"
        if error:
            result["report_reference_errors"].append({"report": finding.report_path, "calgroup": finding.calgroup, "filename": finding.filename, "uid": finding.uid, "message": error})
            continue
        if not 1 <= finding.thimble_id <= 50:
            result["report_reference_errors"].append({"report": finding.report_path, "calgroup": finding.calgroup, "filename": finding.filename, "uid": finding.uid, "message": "RPT Row 超出 1-50"})
            continue
        if not finding.unit_id:
            result["mapping_failures"].append({"group": finding.calgroup, "filename": finding.filename, "row": finding.thimble_id, "message": "SUM Unit 缺失，堆芯位置留空"})
        elif not finding.position:
            result["mapping_failures"].append({"group": finding.calgroup, "filename": finding.filename, "row": finding.thimble_id, "message": "堆芯坐标映射失败"})
        result["findings"].append(finding)
    result["unreported_ect"] = [ect for ect in result["ect"] if not ect["calibration"] and ect["path"].casefold() not in referenced_paths]
    tube_rows: dict[tuple, dict] = {}
    for ect in result["ect"]:
        context = ect["context"]
        if ect["calibration"] or not ect["header_matches_name"] or not 1 <= (ect["row"] or 0) <= 50:
            continue
        key = (context.site_code, context.unit_id, context.outage, ect["row"])
        row = tube_rows.setdefault(key, {"site_code": context.site_code, "unit_id": context.unit_id, "outage": context.outage, "thimble_id": ect["row"], "position": position_for(context.unit_id, ect["row"]) if context.unit_id else "", "entries": set(), "files": set(), "indications": [], "position_groups": set()})
        row["entries"].add(ect["entry"]); row["files"].add(ect["filename"])
    for finding in result["findings"]:
        key = (finding.site_code, finding.unit_id, finding.outage, finding.thimble_id)
        row = tube_rows.setdefault(key, {"site_code": finding.site_code, "unit_id": finding.unit_id, "outage": finding.outage, "thimble_id": finding.thimble_id, "position": finding.position, "entries": set(), "files": set(), "indications": [], "position_groups": set()})
        if finding.entry_no is not None: row["entries"].add(finding.entry_no)
        row["files"].add(finding.filename)
        if is_indication_record(finding):
            row["indications"].append(finding)
            row["position_groups"].add((finding.calgroup.upper(), finding.filename.upper(), finding.location, finding.channel))
    for row in tube_rows.values():
        indications = row.pop("indications")
        row["entry_count"] = len(row.pop("entries")); row["file_count"] = len(row.pop("files"))
        row["raw_indication_count"] = len(indications); row["position_group_count"] = len(row.pop("position_groups"))
        percentages = [finding.percent for finding in indications if finding.percent is not None]
        row["max_percent"] = max(percentages) if percentages else None
    result["tube_summary"] = sorted(tube_rows.values(), key=lambda row: (row["site_code"], row["unit_id"], row["outage"], row["thimble_id"]))
    return result


def report_options(directory: str) -> list[dict]:
    base = Path(directory).expanduser().resolve()
    if not base.is_dir():
        raise ValueError("所选路径不是有效文件夹")
    options = []
    fingerprints: dict[tuple[str, str, str], dict] = {}
    reports = [(parse_sum(group), report) for group in discover_data_groups(base) for report in group.iterdir() if report.is_file() and report.suffix.casefold() == ".rpt"]
    for context, report in sorted(reports, key=lambda item: str(item[1]).casefold()):
        try:
            rows = parse_report(report, context)
            option = {
                "path": str(report),
                "name": report.name,
                "group": report.parent.name,
                "outage": context.outage,
                "site_code": context.site_code,
                "unit_id": context.unit_id,
                "sum_warning": context.warning,
                "analysts": sorted({row.analyst for row in rows if row.analyst}),
                "records": len(rows),
                "tubes": sorted({row.thimble_id for row in rows}),
                "tube_count": len({row.thimble_id for row in rows}),
                "indication_count": sum(is_indication_record(row) for row in rows),
                "blank_indication": blank_indication_summary(rows),
                "duplicate_paths": [],
            }
            digest = hashlib.sha256(report.read_bytes()).hexdigest()
            key = (context.site_code.upper(), context.unit_id, context.component.upper(), context.outage, option["group"].upper(), digest)
            existing = fingerprints.get(key)
            if existing:
                existing["duplicate_paths"].append(str(report))
            else:
                fingerprints[key] = option
                options.append(option)
        except Exception as exc:
            options.append({"path": str(report), "name": report.name, "group": report.parent.name, "error": str(exc), "analysts": [], "records": 0})
    return options


def unique_report_paths(reports) -> list[Path]:
    unique: dict[tuple, Path] = {}
    for report in sorted(reports):
        context = parse_sum(report.parent)
        try:
            key = (context.site_code.upper(), context.unit_id, context.component.upper(), context.outage, report.parent.name.upper(), hashlib.sha256(report.read_bytes()).hexdigest())
        except OSError:
            key = (context.site_code.upper(), context.unit_id, context.component.upper(), context.outage, report.parent.name.upper(), str(report.resolve()).casefold())
        unique.setdefault(key, report)
    return list(unique.values())


def import_directory(directory: str, selected_reports: list[str] | None = None, blank_indication_policy: str = "") -> dict:
    processed = process_directory(directory, selected_reports)
    reports = processed["reports"]
    parsed = inserted = skipped = 0
    errors = list(processed["parse_errors"]) + list(processed["report_reference_errors"])
    blank_indication = blank_indication_summary(processed["findings"])
    if blank_indication["count"] and blank_indication_policy not in {"review", "fill_war"}:
        return {"requires_indication_decision": True, "blank_indication": blank_indication,
                "reports": len(reports), "parsed": len(processed["findings"]), "inserted": 0, "skipped": 0, "errors": errors}
    with connect() as db:
        for finding in processed["findings"]:
            parsed += 1
            raw_columns = asdict(finding)
            columns = indication_columns(finding, blank_indication_policy)
            columns["business_key"] = finding_business_key(raw_columns)
            columns["imported_at"] = datetime.now().isoformat(timespec="seconds")
            keys = list(columns)
            sql = f"INSERT OR IGNORE INTO findings ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})"
            cur = db.execute(sql, [columns[key] for key in keys])
            inserted += cur.rowcount
            skipped += 1 - cur.rowcount
            if not cur.rowcount:
                audit_duplicate(db, columns)
            finding_row = db.execute("SELECT id,indication_resolution FROM findings WHERE source_key=? ORDER BY id LIMIT 1", (columns["source_key"],)).fetchone()
            if finding_row and columns["indication_resolution"] == "bulk_war":
                db.execute("UPDATE findings SET indication='WAR',source_indication='',indication_resolution='bulk_war' WHERE id=?", (finding_row["id"],))
                audit_indication_resolution(db, finding_row["id"], columns)
    return {"reports": len(reports), "parsed": parsed, "inserted": inserted, "skipped": skipped, "errors": errors,
            "blank_indication": blank_indication,
            "raw_report_rows": len(processed["report_rows"]), "tubes": len(processed["tube_summary"]),
            "calibration_ect": len(processed["calibration_ect"]), "unreported_ect": len(processed["unreported_ect"])}


def import_excel_file(filename: str, blank_indication_policy: str = "") -> dict:
    if not blank_indication_policy:
        check = import_excel_file(filename, "check")
        if check["blank_indication"]["count"]:
            check["requires_indication_decision"] = True
            return check
        return import_excel_file(filename, "review")
    dry_run = blank_indication_policy == "check"
    source = Path(filename).expanduser().resolve()
    if not source.is_file() or source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("请选择 .xlsx 或 .xlsm 文件")
    from openpyxl import load_workbook
    book = load_workbook(source, read_only=True, data_only=True)
    aliases = {
        "电站名称": ["电站名称", "基地", "基地名称"], "机组号": ["机组号", "机组", "unitid"],
        "通道编号": ["通道编号", "指套管编号", "incorefluxthimbleid", "套管编号"],
        "在堆芯内的位置": ["在堆芯内的位置", "堆芯位置", "position"], "幅值": ["幅值", "幅值v", "volts"],
        "相位": ["相位", "相位°", "degrees"], "磨损深度": ["磨损深度", "磨损深度%", "percent"],
        "三字码": ["三字码", "缺陷类型", "indication"], "测量通道": ["测量通道", "channel"],
        "磨损位置": ["磨损位置", "location"], "分析人员": ["分析人员", "analyst"],
        "数据": ["数据", "数据文件", "filename"], "数据组": ["数据组", "calgroup"],
        "备注": ["备注", "analysis"], "堵管移位信息": ["堵管移位信息", "管状态"],
        "大修号": ["大修号", "大修", "outage"], "数据流水号": ["数据流水号", "entry", "entryno"],
        "数据点": ["数据点", "datapoint"], "liss区域大小": ["liss区域大小", "lissregionsize"]
    }
    # English templates vary in case, separators and wording. Add normalized aliases
    # by semantic field order while keeping the existing database field names intact.
    # Keep aliases tied to the actual template columns. The previous positional
    # mapping put outage aliases on the unit field and shifted every later field.
    semantic_aliases = {
        0: ["site", "station", "plant", "base"],
        1: ["unit", "unit no", "unit number", "unitid"],
        2: ["channel", "channel no", "channel number", "thimble", "thimble id"],
        3: ["position", "core position", "core location"],
        4: ["volts", "voltage", "amplitude"],
        5: ["phase", "degrees", "angle"],
        6: ["wear depth", "wear", "percent", "wear percent"],
        7: ["indication", "defect", "three character code", "code"],
        8: ["measurement channel", "test channel"],
        9: ["location", "wear location", "defect location"],
        10: ["analyst", "analyst name", "reviewer"],
        11: ["data", "data file", "filename"],
        12: ["data group", "calgroup", "group"],
        13: ["note", "notes", "remark", "remarks"],
        15: ["outage", "overhaul", "overhaul no", "maintenance"],
        16: ["entry", "entry no", "entry number"],
        17: ["datapoint", "data point"],
        18: ["liss region size", "liss size"],
    }
    alias_keys = list(aliases)
    for index, names in semantic_aliases.items():
        if index < len(alias_keys):
            aliases[alias_keys[index]].extend(names)
    normalized_aliases = {key: {normalize_header(item) for item in values} for key, values in aliases.items()}
    selected = None
    for sheet in book.worksheets:
        for header_row in range(1, min(sheet.max_row, 12) + 1):
            headers = [normalize_header(cell.value) for cell in sheet[header_row]]
            mapping = {key: next((index for index, header in enumerate(headers) if header in names), None) for key, names in normalized_aliases.items()}
            if mapping["机组号"] is not None and mapping["通道编号"] is not None:
                selected = (sheet, header_row, mapping)
                break
        if selected:
            break
    if not selected:
        book.close()
        raise ValueError("无法识别Excel表头：至少需要“机组号”和“通道编号/指套管编号”列")
    sheet, header_row, mapping = selected
    rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))
    book.close()
    def cell(values, key):
        index = mapping.get(key)
        return values[index] if index is not None and index < len(values) else None
    inserted = skipped = parsed = 0
    invalid, pending = [], []
    with connect() as db:
        for row_number, values in enumerate(rows, header_row + 1):
            row = {key: cell(values, key) for key in mapping}
            outage, unit_id, thimble_id = outage_from_standard_row(row, source), number(row["机组号"], int), number(row["通道编号"], int)
            if not outage or not unit_id or not thimble_id or not 1 <= thimble_id <= 50:
                if any(value_ not in (None, "") for value_ in values): invalid.append(row_number)
                continue
            parsed += 1
            finding = Finding(
                outage=outage, unit_id=unit_id, thimble_id=thimble_id,
                position=str(row["在堆芯内的位置"] or position_for(unit_id, thimble_id)),
                entry_no=number(row["数据流水号"], int), volts=number(row["幅值"], float), degrees=number(row["相位"], float),
                indication=str(row["三字码"] or ""), percent=number(row["磨损深度"], float), channel=str(row["测量通道"] or ""),
                location=str(row["磨损位置"] or ""), datapoint=number(row["数据点"], int), liss_region_size=number(row["liss区域大小"], int),
                analyst=str(row["分析人员"] or ""), analysis=str(row["备注"] or ""), filepath="",
                filename=str(row["数据"] or source.name), calgroup=str(row["数据组"] or ""), report_path=str(source)
            )
            excel_identity = "|".join(str(value_ or "") for value_ in (
                outage, unit_id, thimble_id, finding.filename, finding.calgroup, finding.location,
                finding.channel, finding.datapoint, finding.analyst, finding.indication, finding.percent,
            ))
            finding.source_key = hashlib.sha256(excel_identity.encode("utf-8")).hexdigest()
            pending.append(finding)
            if dry_run:
                continue
            raw_columns = asdict(finding)
            columns = indication_columns(finding, blank_indication_policy)
            columns["business_key"] = finding_business_key(raw_columns)
            columns["imported_at"] = datetime.now().isoformat(timespec="seconds")
            keys = list(columns)
            sql = f"INSERT OR IGNORE INTO findings ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})"
            cur = db.execute(sql, [columns[key] for key in keys])
            inserted += cur.rowcount
            skipped += 1 - cur.rowcount
            if not cur.rowcount:
                audit_duplicate(db, columns)
            finding_row = db.execute("SELECT id,indication_resolution FROM findings WHERE source_key=? ORDER BY id LIMIT 1", (columns["source_key"],)).fetchone()
            if finding_row and columns["indication_resolution"] == "bulk_war":
                db.execute("UPDATE findings SET indication='WAR',source_indication='',indication_resolution='bulk_war' WHERE id=?", (finding_row["id"],))
                audit_indication_resolution(db, finding_row["id"], columns)
    return {"file": str(source), "sheet": sheet.title, "header_row": header_row, "parsed": parsed,
            "inserted": inserted, "skipped": skipped, "invalid_rows": invalid[:100],
            "blank_indication": blank_indication_summary(pending)}


def analyze_directory(directory: str, selected_reports: list[str] | None = None) -> tuple[list[Finding], list[dict]]:
    processed = process_directory(directory, selected_reports)
    return processed["findings"], processed["parse_errors"] + processed["report_reference_errors"]


def automatic_report_selection(directory: str, policy: str) -> list[str] | None:
    if policy == "all":
        return None
    if policy != "latest":
        return None
    base = Path(directory).expanduser().resolve()
    reports = unique_report_paths(path for group in discover_data_groups(base) for path in group.iterdir() if path.is_file() and path.suffix.casefold() == ".rpt")
    groups: dict[str, list[Path]] = {}
    for report in reports:
        context = parse_sum(report.parent)
        key = f"{context.site_code.upper()}|{context.unit_id}|{context.component.upper()}|{context.outage}|{context.calgroup.upper()}"
        groups.setdefault(key, []).append(report)
    return [str(max(group, key=lambda path: (path.stat().st_mtime, path.name.lower()))) for group in groups.values()]


def analyze_data_groups(directory: str) -> list[dict]:
    groups = []
    for group_dir in discover_data_groups(directory):
        context = parse_sum(group_dir)
        try:
            root = ET.parse(context.sum_path).getroot() if context.sum_path else ET.Element("Summary")
            operator_node = root.find("Operator"); operator_node = operator_node if operator_node is not None else root
            probe_node = root.find("Probe"); probe_node = probe_node if probe_node is not None else root
            valid_ect = [parse_ect_header(path) for path in group_dir.iterdir() if path.is_file() and path.suffix.casefold() == ".ect"]
            valid_ect = [row for row in valid_ect if not row["calibration"]]
            timestamps = []
            for ect in valid_ect:
                if ect["datetime"]: timestamps.append(ect["datetime"])
            groups.append({
                "site_code": context.site_code, "outage": context.outage, "unit_id": context.unit_id,
                "data_group": group_dir.name,
                "operator": value(operator_node, "Id"),
                "probe_type": value(probe_node, "Type"),
                "probe_sn": value(probe_node, "Sn"),
                "probe_model": value(probe_node, "Model"),
                "tube_number": len({row["row"] for row in valid_ect}),
                "entry_number": len(valid_ect),
                "start_time": min(timestamps) if timestamps else "",
                "end_time": max(timestamps) if timestamps else "",
                "report_versions": len([path for path in group_dir.iterdir() if path.is_file() and path.suffix.casefold() == ".rpt"]),
                "warning": context.warning,
            })
        except Exception:
            continue
    return groups


def discover_server_sources(directory: str, max_directories: int = 12000) -> dict:
    base = Path(directory).expanduser()
    if not base.is_dir():
        raise ValueError("服务器根路径不可访问，请检查网络连接和访问权限")
    queue = [(base, 0)]
    visited = 0
    outages: dict[str, dict] = {}
    while queue and visited < max_directories:
        current, depth = queue.pop(0)
        visited += 1
        try:
            children = [path for path in current.iterdir() if path.is_dir()]
        except (OSError, PermissionError):
            continue
        outage_match = re.search(r"(?<![A-Z0-9])([A-Z]\d{3})(?![A-Z0-9])", current.name.upper())
        if outage_match:
            outage = outage_match.group(1)
            local_queue = [(current, 0)]
            thimble_dirs = []
            while local_queue:
                candidate, local_depth = local_queue.pop(0)
                if local_depth > 4:
                    continue
                try:
                    nested = [path for path in candidate.iterdir() if path.is_dir()]
                except (OSError, PermissionError):
                    continue
                for path in nested:
                    try:
                        has_groups = "指套管" in path.name and any(child.is_dir() and child.name.upper().startswith("TH") for child in path.iterdir())
                    except (OSError, PermissionError):
                        has_groups = False
                    if has_groups:
                        thimble_dirs.append(path)
                    elif local_depth < 4:
                        local_queue.append((path, local_depth + 1))
            if thimble_dirs:
                selected = min(thimble_dirs, key=lambda path: len(path.parts))
                station, site_code = infer_station(current, outage)
                unit_folder = current.parent.name
                project_folder = next((part for part in reversed(current.parts[:-1]) if "项目" in part or "市场" in part), current.parent.parent.name if len(current.parents) > 1 else "")
                outages[str(selected).lower()] = {
                    "outage": outage,
                    "unit_id": infer_unit(outage, ""),
                    "site": site_code,
                    "station": station,
                    "path": str(selected),
                    "project": project_folder,
                    "unit_folder": unit_folder,
                    "folder": current.name,
                }
            continue
        if depth < 5:
            queue.extend((path, depth + 1) for path in children)
    items = sorted(outages.values(), key=lambda item: (item["station"], item["unit_id"], item["outage"]), reverse=True)
    return {"root": str(base), "items": items, "scanned_directories": visited, "truncated": bool(queue)}


def export_directory_excel(directory: str, selected_reports: list[str] | None = None, report_policy: str = "manual") -> dict:
    if selected_reports is None and report_policy in {"latest", "all"}:
        selected_reports = automatic_report_selection(directory, report_policy)
    processed = process_directory(directory, selected_reports)
    findings = processed["findings"]
    if not processed["contexts"]:
        raise ValueError("所选目录中没有找到符合 TH数字I数字CAL数字 规则的数据组")
    source_name = Path(directory).resolve().name or "指套管检测数据"
    safe_name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", source_name)
    filename = f"TH_{safe_name}_解析检查表_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "缺陷明细表"
    detail_headers = ["电站名称", "机组号", "安全级别", "设备名称", "被检部位", "通道编号", "在堆芯内的位置", "幅值", "相位", "磨损深度", "三字符", "测量通道", "磨损位置", "探头类型", "分析人员", "数据", "数据组", "备注", "堵管移位信息"]
    sheet.append(detail_headers)
    for finding in findings:
        station = SITE_NAMES.get(finding.site_code.upper(), finding.site_owner or finding.site_code)
        notes = "；".join(value_ for value_ in (finding.comment, finding.analysis, finding.measurement_type) if value_)
        sheet.append([
            station, finding.unit_id or "", "", "", "指套管" if finding.component.upper() == "TH" else finding.component,
            finding.thimble_id, finding.position, finding.volts, finding.degrees, finding.percent,
            finding.indication, finding.channel, finding.location, finding.probe, finding.analyst,
            finding.filename, finding.calgroup, notes, "",
        ])
    header_fill = PatternFill("solid", fgColor="263942")

    tube_sheet = book.create_sheet("管子汇总表")
    tube_sheet.append(["站点代码", "机组号", "大修号", "通道编号", "堆芯位置", "采集入口数", "检测文件数", "原始指示数", "位置组合数", "最大报告百分比"])
    for row in processed["tube_summary"]:
        tube_sheet.append([row["site_code"], row["unit_id"] or "", row["outage"], row["thimble_id"], row["position"], row["entry_count"], row["file_count"], row["raw_indication_count"], row["position_group_count"], row["max_percent"]])

    def ect_sheet(name: str, rows: list[dict]):
        target = book.create_sheet(name)
        target.append(["站点代码", "机组号", "大修号", "数据组", "文件名", "路径", "DIR", "C", "I", "Header Row", "Header Col", "Header Entry", "SHA-256"])
        for row in rows:
            context = row["context"]
            target.append([context.site_code, context.unit_id or "", context.outage, context.calgroup, row["filename"], row["path"], row["name_row"], row["name_col"], row["name_entry"], row["row"], row["col"], row["entry"], row["sha256"]])
        return target

    ect_sheet("未进入RPT的真实ECT", processed["unreported_ect"])
    ect_sheet("DIR999标定文件", processed["calibration_ect"])
    ect_sheet("同名同哈希副本", processed["duplicate_same_hash"])
    ect_sheet("同名不同哈希文件", processed["same_name_different_hash"])

    reference = book.create_sheet("RPT引用异常")
    reference.append(["报告", "数据组", "ECT文件名", "Uid", "异常"])
    for row in processed["report_reference_errors"]:
        reference.append([row["report"], row["calgroup"], row["filename"], row["uid"], row["message"]])

    mismatch = ect_sheet("ECT编号校验异常", processed["ect_mismatches"])
    mismatch.cell(1, 14, "异常")
    for row_index in range(2, mismatch.max_row + 1): mismatch.cell(row_index, 14, "DIR/Row、C/Col 或 I/Entry 不一致")

    missing_sum = book.create_sheet("SUM或机组号缺失")
    missing_sum.append(["数据组", "目录", "SUM", "异常"])
    for row in processed["sum_warnings"]:
        missing_sum.append([row["group"], row["path"], row["sum"], row["message"]])
    for row in processed["parse_errors"]:
        if row["type"].startswith("SUM"): missing_sum.append(["", row["file"], "", row["message"]])

    mapping = book.create_sheet("堆芯映射失败")
    mapping.append(["数据组", "ECT文件名", "通道编号", "异常"])
    for row in processed["mapping_failures"]:
        mapping.append([row["group"], row["filename"], row["row"], row["message"]])

    blank_indication = blank_indication_summary(findings, 10000)
    indication_sheet = book.create_sheet("三字符缺失")
    indication_sheet.append(["大修", "机组", "通道编号", "堆芯位置", "磨损深度", "磨损位置", "测量通道", "数据文件", "数据组", "分析人员", "来源"])
    for row in blank_indication["items"]:
        indication_sheet.append([row["outage"], row["unit_id"], row["thimble_id"], row["position"], row["percent"],
                                 row["location"], row["channel"], row["filename"], row["calgroup"], row["analyst"], row["source"]])

    text_headers = {"在堆芯内的位置", "数据", "数据组", "堆芯位置", "ECT文件名", "文件名"}
    for current in book.worksheets:
        for cell in current[1]:
            cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current.freeze_panes = "A2"
        current.auto_filter.ref = current.dimensions
        current.row_dimensions[1].height = 24
        for column in range(1, current.max_column + 1):
            label = str(current.cell(1, column).value or "")
            if label in text_headers:
                for row_index in range(2, current.max_row + 1): current.cell(row_index, column).number_format = "@"
            values = [str(current.cell(row, column).value or "") for row in range(1, min(current.max_row, 200) + 1)]
            current.column_dimensions[get_column_letter(column)].width = min(56, max(10, len(label) + 3, max(map(len, values), default=0) + 2))
    outcome = workbook_result(filename, book)
    return {**outcome, "rows": len(findings), "raw_report_rows": len(processed["report_rows"]),
            "blank_indication": blank_indication,
            "groups": len(processed["contexts"]), "tubes": len(processed["tube_summary"]),
            "errors": processed["parse_errors"] + processed["report_reference_errors"],
            "unreported_ect": len(processed["unreported_ect"]), "calibration_ect": len(processed["calibration_ect"])}


def query_states() -> dict:
    """Return editable tube-state rows with a best-effort site/position context."""
    with connect() as db:
        rows = db.execute("""
            SELECT s.outage, s.unit_id, s.thimble_id, s.state, s.offset_mm, s.note,
                   COALESCE(MAX(f.site_code), '') AS site_code,
                   COALESCE(MAX(f.position), '') AS position,
                   COALESCE(MAX(f.imported_at), '') AS updated_at
            FROM tube_states s
            LEFT JOIN findings f ON f.outage=s.outage AND f.unit_id=s.unit_id AND f.thimble_id=s.thimble_id
            GROUP BY s.outage, s.unit_id, s.thimble_id, s.state, s.offset_mm, s.note
            ORDER BY s.outage DESC, s.unit_id, s.thimble_id
        """).fetchall()
    return {"items": [dict(row) for row in rows]}


def query_findings(params: dict[str, list[str]]) -> dict:
    page = max(1, number(params.get("page", ["1"])[0], int) or 1)
    size = min(500, max(1, number(params.get("size", ["100"])[0], int) or 100))
    where, args = [], []
    for query_key, column in (("site", "f.site_code"), ("outage", "f.outage"), ("unit", "f.unit_id"), ("analyst", "f.analyst"), ("channel", "f.channel"), ("thimble", "f.thimble_id"), ("calgroup", "f.calgroup")):
        val = params.get(query_key, [""])[0].strip()
        if val:
            where.append(f"{column} = ?" if query_key in {"unit", "thimble"} else f"CAST({column} AS TEXT) LIKE ?")
            args.append(val if query_key in {"unit", "thimble"} else f"%{val}%")
    clause = " WHERE " + " AND ".join(where) if where else ""
    with connect() as db:
        total = db.execute("SELECT COUNT(*) FROM findings f" + clause, args).fetchone()[0]
        sql = """SELECT f.*, COALESCE(s.state,'normal') state, COALESCE(s.offset_mm,0) offset_mm,
                 COALESCE(s.note,'') note FROM findings f LEFT JOIN tube_states s
                 ON s.outage=f.outage AND s.unit_id=f.unit_id AND s.thimble_id=f.thimble_id"""
        rows = db.execute(sql + clause + " ORDER BY f.outage DESC,f.thimble_id,f.entry_no LIMIT ? OFFSET ?", args + [size, (page - 1) * size]).fetchall()
        severity_clause = clause + (" AND " if clause else " WHERE ") + """
            UPPER(TRIM(COALESCE(f.indication,''))) NOT IN ('NDD', 'NONE', 'NO DEFECT')
            AND (TRIM(COALESCE(f.indication,''))<>'' OR TRIM(COALESCE(f.location,''))<>''
                 OR UPPER(TRIM(COALESCE(f.measurement_type,''))) NOT IN ('', 'NONE'))
            AND f.percent IS NOT NULL AND f.percent > 0"""
        core_rows = db.execute("""SELECT f.unit_id, f.thimble_id, f.position,
                MAX(f.percent) AS percent
            FROM findings f""" + severity_clause + """
            GROUP BY f.unit_id, f.thimble_id, f.position
            ORDER BY f.unit_id, f.thimble_id""", args).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        item["location_raw"] = item.get("location", "")
        item["p_zone"], item["p_offset"] = split_location(item["location_raw"])
        item["location"] = format_location(item["location_raw"])
        items.append(item)
    return {"items": items, "core_items": [dict(row) for row in core_rows], "total": total,
            "page": page, "size": size, "pages": max(1, (total + size - 1) // size)}


def query_all_findings(params: dict[str, list[str]]) -> list[dict]:
    """Read every matching row for exports without weakening the UI page limit."""
    first = query_findings({**params, "page": ["1"], "size": ["200"]})
    items = list(first["items"])
    for page in range(2, int(first["pages"]) + 1):
        result = query_findings({**params, "page": [str(page)], "size": ["200"]})
        items.extend(result["items"])
    return items


def overview() -> dict:
    with connect() as db:
        stats = dict(db.execute("SELECT COUNT(*) findings, COUNT(DISTINCT outage) outages, COUNT(DISTINCT unit_id) units, COUNT(DISTINCT analyst) analysts FROM findings").fetchone())
        outages = [row[0] for row in db.execute("SELECT DISTINCT outage FROM findings ORDER BY outage DESC")]
        units = [row[0] for row in db.execute("SELECT DISTINCT unit_id FROM findings ORDER BY unit_id")]
        sites = [row[0] for row in db.execute("SELECT DISTINCT SUBSTR(outage,1,1) FROM findings ORDER BY 1")]
        combinations = [dict(row) for row in db.execute("SELECT DISTINCT SUBSTR(outage,1,1) site, unit_id, outage FROM findings ORDER BY site,unit_id,outage DESC")]
        history = [dict(row) for row in db.execute(
            """SELECT SUBSTR(outage,1,1) site, unit_id, outage,
                      COUNT(*) findings, COUNT(DISTINCT thimble_id) tubes,
                      COUNT(DISTINCT analyst) analysts
                 FROM findings
                GROUP BY outage, unit_id
                ORDER BY site, unit_id, outage"""
        )]
    return {"stats": stats, "sites": sites, "outages": outages, "units": units, "combinations": combinations, "history": history}


def compare(old: str, new: str, unit: int) -> dict:
    if not old or not new or old == new:
        raise ValueError("请选择同一机组的两个不同大修")
    if old[:1].upper() != new[:1].upper() or infer_unit(old, "") != unit or infer_unit(new, "") != unit:
        raise ValueError("仅允许对比相同基地、相同机组的不同大修")
    with connect() as db:
        old_rows = [dict(row) for row in db.execute("SELECT * FROM findings WHERE outage=? AND unit_id=?", (old, unit)).fetchall()]
        new_rows = [dict(row) for row in db.execute("SELECT f.*,COALESCE(s.state,'normal') state,COALESCE(s.offset_mm,0) offset_mm,COALESCE(s.note,'') note FROM findings f LEFT JOIN tube_states s ON s.outage=f.outage AND s.unit_id=f.unit_id AND s.thimble_id=f.thimble_id WHERE f.outage=? AND f.unit_id=?", (new, unit)).fetchall()]
    new_rows = [row for row in new_rows if is_real_defect(row) and row["state"] != "plugged"]
    offsets = {row["thimble_id"]: float(row.get("offset_mm") or 0) for row in new_rows if row["state"] == "shifted"}
    old_by_key = {(row["thimble_id"], *defect_match_key(row, -offsets.get(row["thimble_id"], 0))): row for row in old_rows if is_real_defect(row)}
    items = []
    for row in new_rows:
        key = (row["thimble_id"], *defect_match_key(row))
        item = dict(row)
        previous = old_by_key.get(key)
        item["comparison"] = "NI" if row["state"] == "replaced" or previous is None else "R"
        item["old_volts"] = previous["volts"] if previous else None
        item["old_percent"] = previous["percent"] if previous else None
        item["old_location"] = previous["location"] if previous else ""
        item["old_datapoint"] = previous["datapoint"] if previous else None
        items.append(item)
    return {"old": old, "new": new, "unit": unit, "items": items, "summary": {"R": sum(i["comparison"] == "R" for i in items), "NI": sum(i["comparison"] == "NI" for i in items)}}


def compare_many(outages: list[str], unit: int) -> dict:
    """Compare the latest selected outage with the preceding selected batches."""
    outages = [str(item).strip() for item in outages if str(item).strip()]
    if len(outages) < 2:
        raise ValueError("请选择至少两个不同的大修批次")
    if len(set(outages)) != len(outages):
        raise ValueError("大修批次不能重复")
    if any(infer_unit(outage, "") != unit for outage in outages):
        raise ValueError("仅允许比较同一机组的大修批次")
    if len({outage[:1].upper() for outage in outages}) != 1:
        raise ValueError("仅允许比较同一基地的大修批次")
    with connect() as db:
        all_rows = {}
        for outage in outages:
            rows = db.execute("SELECT f.*,COALESCE(s.state,'normal') state,COALESCE(s.offset_mm,0) offset_mm,COALESCE(s.note,'') note FROM findings f LEFT JOIN tube_states s ON s.outage=f.outage AND s.unit_id=f.unit_id AND s.thimble_id=f.thimble_id WHERE f.outage=? AND f.unit_id=?", (outage, unit)).fetchall()
            all_rows[outage] = [dict(row) for row in rows]
    latest = outages[-1]
    history_by_tube = {}
    for outage in outages:
        for row in all_rows[outage]:
            if is_real_defect(row):
                history_by_tube.setdefault(row["thimble_id"], {}).setdefault(outage, row)
    items = []
    for thimble, history in sorted(history_by_tube.items()):
        current = history.get(latest)
        if not current:
            continue
        previous = None
        for outage in reversed(outages[:-1]):
            if outage in history:
                previous = history[outage]; break
        current_key = defect_match_key(current, float(current.get("offset_mm") or 0) if current.get("state") == "shifted" else 0)
        previous_key = defect_match_key(previous) if previous else None
        comparison = "R" if previous and current_key == previous_key else "NI"
        items.append({"thimble_id": thimble, "position": current.get("position", ""), "comparison": comparison, "note": current.get("note", ""), "history": {outage: {"volts": row.get("volts"), "percent": row.get("percent"), "location": format_location(row.get("location", ""))} for outage, row in history.items()}})
    return {"outages": outages, "new": latest, "unit": unit, "items": items, "summary": {"R": sum(item["comparison"] == "R" for item in items), "NI": sum(item["comparison"] == "NI" for item in items)}}

DATAPOINT_MM = 0.125
COMPARISON_NOTE = "结果比对栏内“R”表示此次显示为原有显示；“NI”表示为此次大修新发现的达到记录标准的显示。"

def normalize_location(location: str) -> tuple[str, float | None]:
    match = re.match(r"^\s*(P\d+)\s*(?:\+\s*([-+]?\d+(?:\.\d+)?))?\s*(?:mm)?\s*$", str(location or ""), re.I)
    return (match.group(1).upper(), number(match.group(2), float)) if match else ((str(location or "").strip(), None))

def is_real_defect(row) -> bool:
    indication = str(row["indication"] or "").strip().upper()
    if indication in {"NDD", "NONE", "NO DEFECT"}:
        return False
    measurement = str(row.get("measurement_type") or "").strip().upper()
    return bool(indication or str(row.get("location") or "").strip() or (measurement and measurement != "NONE"))

def defect_match_key(row, shift_mm: float = 0) -> tuple[str, float]:
    zone, offset = normalize_location(row.get("location", ""))
    if offset is None:
        offset = (row.get("datapoint") or 0) * DATAPOINT_MM
    return zone, round(float(offset) - float(shift_mm or 0), 2)

def annotate_evolution(rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = {}
    for row in rows: grouped.setdefault(row["outage"], []).append(row)
    baseline: set[tuple[str, float]] = set(); blocked = False
    for outage in sorted(grouped, key=outage_sort_key):
        group = grouped[outage]; state = group[0].get("state", "normal"); shift = group[0].get("offset_mm", 0)
        defects = [row for row in group if is_real_defect(row)]
        for row in group:
            row["comparison"] = "堵管后不比对" if blocked else ("无缺陷" if not defects else "")
        if blocked: continue
        if state == "replaced": baseline = set()
        for row in defects:
            row["comparison"] = "R" if baseline and defect_match_key(row, shift if state == "shifted" else 0) in baseline else "NI"
        if state == "plugged":
            blocked = True
            continue
        if defects:
            baseline = {defect_match_key(row) for row in defects}
    return rows

def outage_sort_key(outage: str):
    match = re.search(r"^[A-Z](\d)(\d{2})$", str(outage or "").upper())
    return (int(match.group(1)), int(match.group(2))) if match else (999, 999)

def tube_history(site: str, unit: int, thimble: int) -> dict:
    with connect() as db:
        rows = [dict(row) for row in db.execute("""SELECT f.*, COALESCE(s.state,'normal') state,
            COALESCE(s.offset_mm,0) offset_mm, COALESCE(s.note,'') note
            FROM findings f LEFT JOIN tube_states s ON s.outage=f.outage AND s.unit_id=f.unit_id
            AND s.thimble_id=f.thimble_id WHERE f.unit_id=? AND f.thimble_id=?
            AND SUBSTR(f.outage,1,1)=? ORDER BY f.outage""", (unit, thimble, site.upper()))]
    rows.sort(key=lambda row: outage_sort_key(row["outage"]))
    annotate_evolution(rows)
    for row in rows:
        row["location_raw"] = row.get("location", "")
        row["location"] = format_location(row["location_raw"])
    return {"site": site.upper(), "unit": unit, "thimble": thimble, "items": rows}

def unit_evolution(site: str, unit: int) -> dict:
    with connect() as db:
        thimbles = [row[0] for row in db.execute("SELECT DISTINCT thimble_id FROM findings WHERE unit_id=? AND SUBSTR(outage,1,1)=? ORDER BY thimble_id", (unit, site.upper()))]
    return {"site": site.upper(), "unit": unit, "tubes": [tube_history(site, unit, thimble) for thimble in thimbles]}

def export_evolution_excel(site: str, unit: int) -> dict:
    data = unit_evolution(site, unit)
    book = Workbook(); sheet = book.active; sheet.title = "指套管纵向演变"
    sheet.append(["基地", "机组", "套管", "大修", "状态", "位置", "数据点", "磨损深度", "判定", "备注"])
    for tube in data["tubes"]:
        previous = []
        for row in tube["items"]:
            defects = [item for item in [row] if is_real_defect(item)]
            state = row["state"]
            comparison = row.get("comparison", "无缺陷")
            for item in defects or [row]:
                sheet.append([site, unit, tube["thimble"], row["outage"], state, item["location"], item["datapoint"], item["percent"], comparison, row["note"]])
            if state == "plugged": break
            if defects: previous = defects
    for cell in sheet[1]: cell.font = Font(bold=True)
    return workbook_result(f"TH_{site}_机组{unit}_纵向演变_{datetime.now():%Y%m%d_%H%M%S}.xlsx", book)


def workbook_result(filename: str, book: Workbook) -> dict:
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    target = EXCEL_DIR / filename
    temporary = None
    try:
        with tempfile.NamedTemporaryFile(prefix=f".{target.stem}.", suffix=".tmp", dir=EXCEL_DIR, delete=False) as handle:
            temporary = Path(handle.name)
        book.save(temporary)
        os.replace(temporary, target)
    except Exception:
        if temporary and temporary.exists():
            temporary.unlink(missing_ok=True)
        LOGGER.exception("Excel export failed: %s", target)
        raise
    return {"filename": filename, "file_path": str(target), "download_url": f"/api/download-excel?name={quote(filename)}"}


def export_comparison_excel(old: str, new: str, unit: int) -> dict:
    data = compare(old, new, unit)
    book = Workbook()
    sheet = book.active
    sheet.title = "涡流检验结果对比表"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = f"{unit}号机组反应堆中子通量测量指套管涡流检验结果对比表"
    sheet["A1"].font = Font(size=15, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["序号", "通道编号", "堆芯位置", new, "", "", old, "", "", "备注"])
    sheet.append(["", "", "", "幅值（V）", "磨损深度（壁厚%）", "磨损位置（mm）", "幅值（V）", "磨损深度（壁厚%）", "磨损位置（mm）", ""])
    for column in ("A", "B", "C", "J"):
        sheet.merge_cells(f"{column}2:{column}3")
    sheet.merge_cells("D2:F2")
    sheet.merge_cells("G2:I2")
    for index, row in enumerate(data["items"], 1):
        sheet.append([
            index, row["thimble_id"], row["position"],
            "/" if row.get("volts") is None else row["volts"],
            "/" if row.get("percent") is None else row["percent"],
            format_location(row.get("location") or "") or "/",
            "/" if row.get("old_volts") is None else row["old_volts"],
            "/" if row.get("old_percent") is None else row["old_percent"],
            format_location(row.get("old_location") or "") or "/",
            row["comparison"],
        ])
    spans = report_tube_rowspans(data["items"])
    for offset, span in enumerate(spans, 4):
        if span > 1:
            sheet.merge_cells(start_row=offset, end_row=offset + span - 1, start_column=2, end_column=2)
            sheet.merge_cells(start_row=offset, end_row=offset + span - 1, start_column=3, end_column=3)
    note_row = sheet.max_row + 1
    sheet.merge_cells(start_row=note_row, end_row=note_row, start_column=1, end_column=10)
    sheet.cell(note_row, 1, f"备注：{COMPARISON_NOTE}")
    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="DCE6EA")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A4"
    for column in range(1, 11):
        sheet.column_dimensions[get_column_letter(column)].width = 16 if 4 <= column <= 9 else 11
    return {**workbook_result(f"TH_机组{unit}_{new}_对比_{old}_演变对比_{datetime.now():%Y%m%d_%H%M%S}.xlsx", book), **data["summary"]}


def export_comparison_excel_many(outages: list[str], unit: int) -> dict:
    data = compare_many(outages, unit)
    book = Workbook(); sheet = book.active; sheet.title = "多次大修对比"
    headers = ["判定", "套管", "堆芯位置"]
    for outage in data["outages"]:
        headers.extend([f"{outage}幅值(V)", f"{outage}磨损深度(%)", f"{outage}磨损位置"])
    headers.append("备注"); sheet.append(headers)
    for row in data["items"]:
        values = [row["comparison"], row["thimble_id"], row["position"]]
        for outage in data["outages"]:
            item = row.get("history", {}).get(outage, {})
            values.extend([item.get("volts", ""), item.get("percent", ""), item.get("location", "")])
        values.append(row.get("note", "")); sheet.append(values)
    for cell in sheet[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    for column in range(1, len(headers) + 1): sheet.column_dimensions[get_column_letter(column)].width = 16
    return {**workbook_result(f"TH_机组{unit}_{data['new']}_多次大修对比_{datetime.now():%Y%m%d_%H%M%S}.xlsx", book), **data["summary"]}


def export_inspection_report(outage: str, unit: int, metadata: dict) -> dict:
    if infer_unit(outage, "") != unit:
        raise ValueError("大修编号与机组不匹配")
    with connect() as db:
        rows = [dict(row) for row in db.execute("SELECT * FROM findings WHERE outage=? AND unit_id=? ORDER BY thimble_id,entry_no,datapoint", (outage, unit))]
    if not rows:
        raise ValueError("当前条件没有可生成报告的数据")
    book = Workbook()
    sheet = book.active
    sheet.title = "涡流检验报告单"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = metadata.get("title") or "反应堆中子通量测量指套管涡流检验报告单"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    info = [
        ("电厂名称", metadata.get("plant", "")), ("机组", unit), ("检查类型", metadata.get("inspection_type", outage)),
        ("设备/部件名称", metadata.get("component", "指套管")), ("安全等级", metadata.get("safety_class", "")), ("方向号", metadata.get("direction", "")),
        ("材料", metadata.get("material", "")), ("尺寸", metadata.get("size", "")), ("报告单编号", metadata.get("report_no", "")),
    ]
    for index in range(0, len(info), 3):
        row = 3 + index // 3
        for offset, (label, value_) in enumerate(info[index:index + 3]):
            col = 1 + offset * 2
            sheet.cell(row, col, label).font = Font(bold=True)
            sheet.cell(row, col + 1, value_)
    header_row = 7
    headers = ["序号", "通道编号", "堆芯位置", "显示类型", "幅值(V)", "磨损深度(壁厚%)", "磨损位置", "测量通道"]
    for col, label in enumerate(headers, 1):
        sheet.cell(header_row, col, label)
    for index, row in enumerate(rows, 1):
        display = display_indication(row)
        values = [index, row["thimble_id"], row["position"], display, row["volts"], row["percent"], row["location"], row["channel"]]
        for col, value_ in enumerate(values, 1):
            sheet.cell(header_row + index, col, value_)
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor="DCE6EA")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    widths = [8, 12, 13, 12, 12, 18, 18, 16]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    return {**workbook_result(f"TH_机组{unit}_{outage}_检验结果_{datetime.now():%Y%m%d_%H%M%S}.xlsx", book), "rows": len(rows)}


def inspection_report_rows(outage: str, unit: int, finding_ids: list[int] | None = None) -> list[dict]:
    outage = str(outage or "").strip().upper()
    if not outage or not unit:
        raise ValueError("请选择大修和机组")
    ids = sorted({int(value) for value in (finding_ids or []) if str(value).isdigit() and int(value) > 0})
    sql = "SELECT * FROM findings WHERE outage=? AND unit_id=?"
    args: list = [outage, unit]
    if ids:
        sql += f" AND id IN ({','.join('?' for _ in ids)})"
        args.extend(ids)
    sql += " ORDER BY thimble_id,entry_no,datapoint,id"
    with connect() as db:
        rows = [dict(row) for row in db.execute(sql, args)]
    if not rows:
        raise ValueError("当前条件或勾选记录没有可生成报告的数据")
    return rows


def report_tube_rowspans(rows: list[dict]) -> list[int]:
    """Return per-row spans for consecutive records of the same physical tube."""
    spans = [1] * len(rows)
    index = 0
    while index < len(rows):
        channel = rows[index].get("thimble_id")
        position = str(rows[index].get("position") or "").strip()
        end = index + 1
        if channel not in (None, "") and position:
            while end < len(rows):
                candidate = rows[end]
                if candidate.get("thimble_id") != channel or str(candidate.get("position") or "").strip() != position:
                    break
                end += 1
        if end - index > 1:
            spans[index] = end - index
            for continuation in range(index + 1, end):
                spans[continuation] = 0
        index = end
    return spans


def report_metadata(metadata: dict, outage: str, unit: int) -> list[tuple[str, str]]:
    return [
        ("设备/部件编号", str(metadata.get("component_no") or "")),
        ("涡流检验报告单（TH）", ""),
        ("报告单编号", str(metadata.get("report_no") or "")),
        ("电厂名称", str(metadata.get("plant") or "")),
        ("机组", str(unit)),
        ("检查类型", str(metadata.get("inspection") or outage)),
        ("设备/部件名称", str(metadata.get("component") or "指套管")),
        ("安全等级", str(metadata.get("safety") or "")),
        ("房间号", str(metadata.get("room") or "")),
        ("检验范围", str(metadata.get("scope") or "")),
        ("材料", str(metadata.get("material") or "")),
        ("尺寸", str(metadata.get("size") or "Φ8.6 × 1.7 mm")),
        ("涡流仪型号", str(metadata.get("instrument_model") or "")),
        ("涡流仪出厂编号", str(metadata.get("instrument_serial") or "")),
        ("标定管编号", str(metadata.get("calibration_no") or "")),
        ("探头类型", str(metadata.get("probe_type") or "")),
        ("探头型号", str(metadata.get("probe_model") or "")),
        ("探头规格", str(metadata.get("probe_spec") or "Φ4.8 mm")),
        ("检验速度", str(metadata.get("speed") or "250 mm/s")),
        ("采样率", str(metadata.get("sample_rate") or "2000 点/秒")),
        ("检验频率", str(metadata.get("frequency") or "160、80、40、20 kHz")),
        ("检验程序/版次", str(metadata.get("procedure") or "")),
    ]


def enrich_report_metadata(metadata: dict, rows: list[dict], outage: str, unit: int) -> dict:
    """Fill only values that are present in imported SUM/RPT data; never guess identifiers."""
    result = dict(metadata or {})
    first = rows[0] if rows else {}
    defaults = {
        "plant": first.get("site_owner") or first.get("site_code") or "",
        "inspection": outage,
        "component": "指套管",
        "instrument_model": first.get("tester") or "",
        "probe_model": first.get("probe") or "",
        "probe_type": "BOBBIN",
        "probe_spec": "Φ4.8 mm",
        "size": "Φ8.6 × 1.7 mm",
        "speed": "250 mm/s",
        "sample_rate": "2000 点/秒",
        "frequency": "160、80、40、20 kHz",
    }
    for key, value in defaults.items():
        if not str(result.get(key) or "").strip() and value:
            result[key] = value
    result["unit"] = unit
    return result


def _w_text(value) -> str:
    return html.escape("" if value is None else str(value), quote=False).replace("\n", "<w:br/>")


def _w_paragraph(value="", align="center", bold=False, size=20) -> str:
    alignment = f'<w:jc w:val="{align}"/>' if align else ""
    weight = "<w:b/>" if bold else ""
    return f'<w:p><w:pPr>{alignment}</w:pPr><w:r><w:rPr>{weight}<w:sz w:val="{size}"/><w:rFonts w:ascii="SimSun" w:hAnsi="SimSun" w:eastAsia="宋体"/></w:rPr><w:t xml:space="preserve">{_w_text(value)}</w:t></w:r></w:p>'


def _w_table(headers: list[str], rows: list[list], merge_group_columns: tuple[int, ...] = (), merge_columns: tuple[int, ...] = ()) -> str:
    count = len(headers); width = max(900, 10200 // max(1, count))
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for _ in headers)
    merge_states = ["none"] * len(rows)
    if merge_group_columns and merge_columns:
        groups = [tuple(row[column] for column in merge_group_columns) for row in rows]
        for index, group in enumerate(groups):
            if any(value in (None, "") for value in group):
                continue
            previous_same = index > 0 and groups[index - 1] == group
            next_same = index + 1 < len(groups) and groups[index + 1] == group
            if previous_same:
                merge_states[index] = "continue"
            elif next_same:
                merge_states[index] = "restart"
    def row_xml(values, header=False, merge_state="none"):
        cells = []
        for column, value in enumerate(values):
            vertical_merge = ""
            if not header and column in merge_columns and merge_state != "none":
                vertical_merge = '<w:vMerge w:val="restart"/>' if merge_state == "restart" else "<w:vMerge/>"
                if merge_state == "continue":
                    value = ""
            cells.append(f'<w:tc><w:tcPr><w:tcW w:w="{width}" w:type="dxa"/><w:vAlign w:val="center"/>{vertical_merge}</w:tcPr>{_w_paragraph(value, "center", header, 16)}</w:tc>')
        return f'<w:tr>{"".join(cells)}</w:tr>'
    body = "".join(row_xml(row, merge_state=merge_states[index]) for index, row in enumerate(rows))
    return f'<w:tbl><w:tblPr><w:tblBorders><w:top w:val="single" w:sz="6"/><w:left w:val="single" w:sz="6"/><w:bottom w:val="single" w:sz="6"/><w:right w:val="single" w:sz="6"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders><w:tblLayout w:type="fixed"/></w:tblPr><w:tblGrid>{grid}</w:tblGrid>{row_xml(headers, True)}{body}</w:tbl>'


def _w_comparison_table(new: str, old: str, rows: list[list]) -> str:
    """Build the formal two-level outage comparison table."""
    widths = [520, 760, 820, 800, 900, 980, 800, 900, 980, 560]
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)

    def cell(value, width, *, bold=False, grid_span=1, vmerge=""):
        span = f'<w:gridSpan w:val="{grid_span}"/>' if grid_span > 1 else ""
        merge = '<w:vMerge w:val="restart"/>' if vmerge == "restart" else ("<w:vMerge/>" if vmerge == "continue" else "")
        return f'<w:tc><w:tcPr><w:tcW w:w="{width}" w:type="dxa"/><w:vAlign w:val="center"/>{span}{merge}</w:tcPr>{_w_paragraph(value, "center", bold, 15)}</w:tc>'

    header_one = "".join([
        cell("序号", widths[0], bold=True, vmerge="restart"),
        cell("通道编号", widths[1], bold=True, vmerge="restart"),
        cell("堆芯位置", widths[2], bold=True, vmerge="restart"),
        cell(new, sum(widths[3:6]), bold=True, grid_span=3),
        cell(old, sum(widths[6:9]), bold=True, grid_span=3),
        cell("备注", widths[9], bold=True, vmerge="restart"),
    ])
    subheaders = ["幅值（V）", "磨损深度（壁厚%）", "磨损位置（mm）"]
    header_two = "".join([
        cell("", widths[0], vmerge="continue"),
        cell("", widths[1], vmerge="continue"),
        cell("", widths[2], vmerge="continue"),
        *(cell(label, widths[3 + index], bold=True) for index, label in enumerate(subheaders)),
        *(cell(label, widths[6 + index], bold=True) for index, label in enumerate(subheaders)),
        cell("", widths[9], vmerge="continue"),
    ])

    groups = [(row[1], str(row[2] or "").strip()) for row in rows]
    body_rows = []
    for row_index, row in enumerate(rows):
        group = groups[row_index]
        merge_state = "none"
        if group[0] not in (None, "") and group[1]:
            if row_index > 0 and groups[row_index - 1] == group:
                merge_state = "continue"
            elif row_index + 1 < len(groups) and groups[row_index + 1] == group:
                merge_state = "restart"
        body_cells = []
        for column, value in enumerate(row):
            state = merge_state if column in (1, 2) else ""
            if state == "continue":
                value = ""
            body_cells.append(cell(value, widths[column], vmerge=state))
        body_rows.append(f'<w:tr>{"".join(body_cells)}</w:tr>')

    borders = '<w:tblBorders><w:top w:val="single" w:sz="6"/><w:left w:val="single" w:sz="6"/><w:bottom w:val="single" w:sz="6"/><w:right w:val="single" w:sz="6"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders>'
    return f'<w:tbl><w:tblPr>{borders}<w:tblLayout w:type="fixed"/></w:tblPr><w:tblGrid>{grid}</w:tblGrid><w:tr>{header_one}</w:tr><w:tr>{header_two}</w:tr>{"".join(body_rows)}</w:tbl>'


def _w_report_info_table(pairs: list[tuple[str, str]]) -> str:
    """Build the formal three-column label/value header used by the TH report sheet."""
    widths = [1050, 1800, 1050, 1800, 1050, 1800]
    grid = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)
    cells = []
    for index in range(0, len(pairs), 3):
        row = pairs[index:index + 3]
        row += [("", "")] * (3 - len(row))
        row_cells = []
        for pair_index, (label, value) in enumerate(row):
            label_width, value_width = widths[pair_index * 2:pair_index * 2 + 2]
            row_cells.append(
                f'<w:tc><w:tcPr><w:tcW w:w="{label_width}" w:type="dxa"/><w:vAlign w:val="center"/></w:tcPr>{_w_paragraph(label, "center", True, 15)}</w:tc>'
            )
            row_cells.append(
                f'<w:tc><w:tcPr><w:tcW w:w="{value_width}" w:type="dxa"/><w:vAlign w:val="center"/></w:tcPr>{_w_paragraph(value, "center", False, 15)}</w:tc>'
            )
        cells.append(f'<w:tr>{"".join(row_cells)}</w:tr>')
    return f'<w:tbl><w:tblPr><w:tblBorders><w:top w:val="single" w:sz="6"/><w:left w:val="single" w:sz="6"/><w:bottom w:val="single" w:sz="6"/><w:right w:val="single" w:sz="6"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders><w:tblLayout w:type="fixed"/></w:tblPr><w:tblGrid>{grid}</w:tblGrid>{"".join(cells)}</w:tbl>'


def _w_page_break() -> str:
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def _write_docx(filename: str, blocks: list[str], landscape=True) -> dict:
    from zipfile import ZIP_DEFLATED, ZipFile
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    target = REPORT_DIR / filename; temporary = target.with_name(f".{target.stem}.{os.getpid()}.tmp")
    size = '<w:pgSz w:w="15840" w:h="12240"/>' if landscape else '<w:pgSz w:w="12240" w:h="15840"/>'
    document_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>' + ''.join(blocks) + f'<w:sectPr>{size}<w:pgMar w:top="792" w:right="792" w:bottom="792" w:left="792"/></w:sectPr></w:body></w:document>'
    content_types = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/></Types>'
    rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>'
    word_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with ZipFile(temporary, "w", ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types); archive.writestr("_rels/.rels", rels); archive.writestr("word/document.xml", document_xml); archive.writestr("word/_rels/document.xml.rels", word_rels)
        os.replace(temporary, target)
    except Exception:
        temporary.unlink(missing_ok=True); LOGGER.exception("Word report export failed: %s", target); raise
    return {"filename": filename, "file_path": str(target), "download_url": f"/api/download-report?name={quote(filename)}"}


def build_inspection_preview(outage: str, unit: int, metadata: dict, finding_ids: list[int] | None = None) -> dict:
    rows = inspection_report_rows(outage, unit, finding_ids)
    metadata = enrich_report_metadata(metadata, rows, outage, unit)
    title = metadata.get("title") or "反应堆中子通量测量指套管涡流检验报告单"
    info = "".join(f"<div><span>{html.escape(label)}</span><b>{html.escape(value or '待填写')}</b></div>" for label, value in report_metadata(metadata, outage, unit))
    body_rows = []
    for index, (row, rowspan) in enumerate(zip(rows, report_tube_rowspans(rows)), 1):
        tube_cells = ""
        if rowspan:
            span = f" rowspan='{rowspan}'" if rowspan > 1 else ""
            tube_cells = f"<td{span}>{row['thimble_id']}</td><td{span}>{html.escape(row.get('position') or '')}</td>"
        body_rows.append(
            f"<tr><td>{index}</td>{tube_cells}<td>{html.escape(display_indication(row))}</td>"
            f"<td>{'/' if row.get('volts') is None else row['volts']}</td>"
            f"<td>{'/' if row.get('percent') is None else row['percent']}</td>"
            f"<td>{html.escape('/' if display_indication(row) == 'NDD' else format_location(row.get('location') or ''))}</td>"
            f"<td>{html.escape('/' if display_indication(row) == 'NDD' else row.get('channel') or '')}</td></tr>"
        )
    body = "".join(body_rows)
    preview = f"<article class='report-sheet'><div class='report-page-mark'>II-1/1</div><h1>{html.escape(title)}</h1><section class='report-info'>{info}</section><h2>检验结果</h2><div class='report-table-wrap'><table><thead><tr><th>序号</th><th>通道编号</th><th>堆芯位置</th><th>显示类型</th><th>幅值（V）</th><th>磨损深度（壁厚%）</th><th>磨损位置</th><th>测量通道</th></tr></thead><tbody>{body}</tbody></table></div><footer>备注：无。<br><br>分析人员/级别：____________　审核/级别：____________　批准：____________　业主：____________</footer></article>"
    return {"title": title, "rows": len(rows), "html": preview}


def export_inspection_docx(outage: str, unit: int, metadata: dict, finding_ids: list[int] | None = None) -> dict:
    rows = inspection_report_rows(outage, unit, finding_ids)
    metadata = enrich_report_metadata(metadata, rows, outage, unit)
    title = metadata.get("title") or "反应堆中子通量测量指套管涡流检验报告单"
    columns = ["序号", "通道编号", "堆芯位置", "显示类型", "幅值（V）", "磨损深度（壁厚%）", "磨损位置", "测量通道"]
    pages = [rows[index:index + 20] for index in range(0, len(rows), 20)] or [[]]
    blocks = []
    for page_index, page_rows in enumerate(pages, 1):
        blocks.append(_w_paragraph(f"II-{page_index}/{len(pages)}", "right", False, 14))
        blocks.append(_w_paragraph(title, "center", True, 24))
        blocks.append(_w_report_info_table(report_metadata(metadata, outage, unit)))
        blocks.append(_w_paragraph("检验结果", "center", True, 18))
        values = []
        for row_index, row in enumerate(page_rows, page_index * 20 - 19):
            indication = display_indication(row)
            is_ndd = str(indication).strip().upper() in {"NDD", "NONE", "NO DEFECT"}
            values.append([
                row_index, row.get("thimble_id") or "", row.get("position") or "", indication,
                "/" if is_ndd or row.get("volts") is None else row.get("volts"),
                "/" if is_ndd or row.get("percent") is None else row.get("percent"),
                "/" if is_ndd else format_location(row.get("location") or ""),
                "/" if is_ndd else row.get("channel") or "",
            ])
        blocks.append(_w_table(columns, values, merge_group_columns=(1, 2), merge_columns=(1, 2)))
        if page_index == len(pages):
            blocks.append(_w_paragraph("备注：\n无。", "left", False, 15))
            blocks.append(_w_table(["", "分析人员/级别", "审核/级别", "批准", "业主"], [["签名", "", "", "", ""], ["日期", "", "", "", ""]]))
        if page_index < len(pages): blocks.append(_w_page_break())
    return {**_write_docx(f"TH_机组{unit}_{outage}_检验结果_{datetime.now():%Y%m%d_%H%M%S}.docx", blocks, landscape=False), "rows": len(rows)}


def build_comparison_preview(old: str, new: str, unit: int) -> dict:
    data = compare(old, new, unit)
    body_rows = []
    for index, (row, rowspan) in enumerate(zip(data["items"], report_tube_rowspans(data["items"])), 1):
        tube_cells = ""
        if rowspan:
            span = f" rowspan='{rowspan}'" if rowspan > 1 else ""
            tube_cells = f"<td{span}>{row['thimble_id']}</td><td{span}>{html.escape(row.get('position') or '')}</td>"
        old_values = (
            '/' if row.get('old_volts') is None else row['old_volts'],
            '/' if row.get('old_percent') is None else row['old_percent'],
            '/' if not row.get('old_location') else format_location(row['old_location']),
        )
        body_rows.append(
            f"<tr><td>{index}</td>{tube_cells}"
            f"<td>{'' if row.get('volts') is None else row['volts']}</td><td>{'' if row.get('percent') is None else row['percent']}</td><td>{html.escape(format_location(row.get('location') or ''))}</td>"
            f"<td>{old_values[0]}</td><td>{old_values[1]}</td><td>{html.escape(str(old_values[2]))}</td><td><b class='{row['comparison'].lower()}'>{row['comparison']}</b></td></tr>"
        )
    body = "".join(body_rows)
    title = f"{unit}号机组反应堆中子通量测量指套管涡流检验结果对比表"
    preview = f"<article class='report-sheet comparison-sheet'><h1>{html.escape(title)}</h1><p class='report-lead'>本次大修（{html.escape(new)}）与历史大修（{html.escape(old)}）结果对比。历史缺陷 R：{data['summary']['R']}，新增缺陷 NI：{data['summary']['NI']}。</p><div class='report-table-wrap'><table><caption>表 5　涡流检验结果比对表</caption><thead><tr><th rowspan='2'>序号</th><th rowspan='2'>通道编号</th><th rowspan='2'>堆芯位置</th><th colspan='3'>{html.escape(new)}</th><th colspan='3'>{html.escape(old)}</th><th rowspan='2'>结果比对</th></tr><tr><th>幅值（V）</th><th>磨损深度（壁厚%）</th><th>磨损位置（mm）</th><th>幅值（V）</th><th>磨损深度（壁厚%）</th><th>磨损位置（mm）</th></tr></thead><tbody>{body}</tbody></table></div><footer>备注：{html.escape(COMPARISON_NOTE)}</footer></article>"
    return {"title": title, "rows": len(data["items"]), "html": preview, **data["summary"]}


def export_comparison_docx(old: str, new: str, unit: int) -> dict:
    data = compare(old, new, unit)
    values = []
    for index, row in enumerate(data["items"], 1):
        values.append([
            index, row["thimble_id"], row.get("position") or "",
            "/" if row.get("volts") is None else row.get("volts"),
            "/" if row.get("percent") is None else row.get("percent"),
            format_location(row.get("location") or "") or "/",
            "/" if row.get("old_volts") is None else row.get("old_volts"),
            "/" if row.get("old_percent") is None else row.get("old_percent"),
            format_location(row.get("old_location") or "") or "/",
            row["comparison"],
        ])
    pages = [values[index:index + 25] for index in range(0, len(values), 25)] or [[]]
    blocks = []
    for page_index, page_rows in enumerate(pages, 1):
        blocks.append(_w_paragraph(f"II-{page_index}/{len(pages)}", "right", False, 14))
        blocks.append(_w_paragraph(f"{unit}号机组反应堆中子通量测量指套管涡流检验结果对比表", "center", True, 24))
        blocks.append(_w_paragraph(f"本次大修（{new}）与历史大修（{old}）的结果对比如下。", "left", False, 17))
        blocks.append(_w_paragraph("表 5　涡流检验结果比对表", "center", True, 17))
        blocks.append(_w_comparison_table(new, old, page_rows))
        if page_index == len(pages):
            blocks.append(_w_paragraph(f"备注：{COMPARISON_NOTE}", "left", False, 15))
        if page_index < len(pages):
            blocks.append(_w_page_break())
    return {**_write_docx(f"TH_机组{unit}_{new}_对比_{old}_综合报告_{datetime.now():%Y%m%d_%H%M%S}.docx", blocks, landscape=False), "rows": len(data["items"]), **data["summary"]}


def health_status() -> dict:
    return {
        "ok": True,
        "service": "thimble-local",
        "version": SERVICE_VERSION,
        "db_path": str(DB_PATH),
        "output_dir": str(EXCEL_DIR),
        "database_exists": DB_PATH.is_file(),
    }


def validate_state_payload(data: dict) -> tuple[str, int, int, str, float, str]:
    outage = str(data.get("outage", "")).strip().upper()
    unit = number(data.get("unit_id"), int)
    thimble = number(data.get("thimble_id"), int)
    state = str(data.get("state", "")).strip().lower()
    offset = number(data.get("offset_mm", 0), float) or 0.0
    note = str(data.get("note", "")).strip()
    if not re.fullmatch(r"[A-Z]\d{3}", outage):
        raise ValueError("大修编号格式应为基地字母加三位数字")
    if not unit or unit < 1 or thimble < 1 or thimble > 50:
        raise ValueError("机组号或套管号无效")
    if state not in {"normal", "plugged", "replaced", "shifted"}:
        raise ValueError("管状态无效")
    if offset < 0 or offset > 1000:
        raise ValueError("位移量应在 0 至 1000 mm 之间")
    return outage, unit, thimble, state, offset, note


class Handler(SimpleHTTPRequestHandler):
    MAX_BODY_BYTES = 8 * 1024 * 1024

    def log_message(self, fmt, *args):
        sys.stdout.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def json_response(self, data, status=200):
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def body(self):
        length = int(self.headers.get("Content-Length", 0) or 0)
        if length > self.MAX_BODY_BYTES:
            raise ValueError("请求体过大，请分批导入文件")
        try:
            return json.loads(self.rfile.read(length) or b"{}")
        except json.JSONDecodeError as exc:
            raise ValueError("请求数据格式无效") from exc

    def do_GET(self):
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/health": return self.json_response(health_status())
            if parsed.path == "/api/overview": return self.json_response(overview())
            if parsed.path == "/api/duplicates": return self.json_response(duplicate_summary())
            if parsed.path == "/api/states": return self.json_response(query_states())
            if parsed.path == "/api/findings": return self.json_response(query_findings(parse_qs(parsed.query)))
            if parsed.path == "/api/tube-history":
                q = parse_qs(parsed.query); return self.json_response(tube_history(q.get("site", [""])[0], int(q.get("unit", ["0"])[0]), int(q.get("thimble", ["0"])[0])))
            if parsed.path == "/api/unit-evolution":
                q = parse_qs(parsed.query); return self.json_response(unit_evolution(q.get("site", [""])[0], int(q.get("unit", ["0"])[0])))
            if parsed.path == "/api/compare":
                q = parse_qs(parsed.query)
                if q.get("outages"):
                    return self.json_response(compare_many(q.get("outages", [""])[0].split(","), int(q.get("unit", ["0"])[0])))
                return self.json_response(compare(q.get("old", [""])[0], q.get("new", [""])[0], int(q.get("unit", ["0"])[0])))
            if parsed.path == "/api/export": return self.export_csv(parse_qs(parsed.query))
            if parsed.path == "/api/download-excel": return self.download_excel(parse_qs(parsed.query))
            if parsed.path == "/api/download-report": return self.download_report(parse_qs(parsed.query))
            return self.serve_static(parsed.path)
        except Exception as exc:
            LOGGER.exception("GET %s failed", self.path)
            self.json_response({"error": str(exc) if isinstance(exc, ValueError) else "本地服务处理失败，请查看日志"}, 400 if isinstance(exc, ValueError) else 500)

    def do_POST(self):
        try:
            data = self.body()
            if self.path == "/api/report-options": return self.json_response({"reports": report_options(data.get("path", ""))})
            if self.path == "/api/discover-server": return self.json_response(discover_server_sources(data.get("path", "")))
            if self.path == "/api/import": return self.json_response(import_directory(data.get("path", ""), data.get("reports"), data.get("blank_indication_policy", "")))
            if self.path == "/api/import-excel": return self.json_response(import_excel_file(data.get("path", ""), data.get("blank_indication_policy", "")))
            if self.path == "/api/deduplicate": return self.json_response(deduplicate_database())
            if self.path == "/api/export-comparison":
                outages = data.get("outages") or []
                if len(outages) >= 2:
                    return self.json_response(export_comparison_docx(outages[-2], outages[-1], int(data.get("unit", 0))))
                return self.json_response(export_comparison_docx(data.get("old", ""), data.get("new", ""), int(data.get("unit", 0))))
            if self.path == "/api/export-evolution": return self.json_response(export_evolution_excel(data.get("site", ""), int(data.get("unit", 0))))
            if self.path == "/api/export-report": return self.json_response(export_inspection_docx(data.get("outage", ""), int(data.get("unit", 0)), data.get("metadata") or {}, data.get("finding_ids") or []))
            if self.path == "/api/report-preview":
                if data.get("kind") == "comparison":
                    return self.json_response(build_comparison_preview(data.get("old", ""), data.get("new", ""), int(data.get("unit", 0))))
                return self.json_response(build_inspection_preview(data.get("outage", ""), int(data.get("unit", 0)), data.get("metadata") or {}, data.get("finding_ids") or []))
            if self.path == "/api/clear":
                if data.get("confirmed") is not True:
                    return self.json_response({"error": "清空操作必须明确确认"}, 400)
                with connect() as db:
                    db.execute("DELETE FROM findings")
                    db.execute("DELETE FROM tube_states")
                return self.json_response({"ok": True})
            if self.path == "/api/export-folder-excel": return self.json_response(export_directory_excel(data.get("path", ""), data.get("reports"), data.get("report_policy", "manual")))
            if self.path == "/api/state":
                outage, unit, thimble, state, offset, note = validate_state_payload(data)
                with connect() as db:
                    db.execute("INSERT INTO tube_states(outage,unit_id,thimble_id,state,offset_mm,note) VALUES(?,?,?,?,?,?) ON CONFLICT(outage,unit_id,thimble_id) DO UPDATE SET state=excluded.state,offset_mm=excluded.offset_mm,note=excluded.note", (outage, unit, thimble, state, offset, note))
                return self.json_response({"ok": True})
            if self.path == "/api/state-delete":
                outage = str(data.get("outage", "")).strip()
                unit = number(data.get("unit_id"), int)
                thimble = number(data.get("thimble_id"), int)
                if not outage or not unit or not thimble:
                    raise ValueError("删除管状态需要大修、机组和通道编号")
                with connect() as db:
                    db.execute("DELETE FROM tube_states WHERE outage=? AND unit_id=? AND thimble_id=?", (outage, unit, thimble))
                return self.json_response({"ok": True})
            self.json_response({"error": "接口不存在"}, 404)
        except Exception as exc:
            LOGGER.exception("POST %s failed", self.path)
            self.json_response({"error": str(exc) if isinstance(exc, ValueError) else "本地服务处理失败，请查看日志"}, 400 if isinstance(exc, ValueError) else 500)

    def export_csv(self, params):
        rows = query_all_findings(params)
        output = io.StringIO()
        fields = ["outage","unit_id","thimble_id","position","entry_no","volts","degrees","indication","percent","channel","location","datapoint","analyst","analysis","filename","calgroup","state","offset_mm","note"]
        writer = csv.DictWriter(output, fields, extrasaction="ignore")
        writer.writeheader(); writer.writerows(rows)
        payload = b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
        self.send_response(200); self.send_header("Content-Type", "text/csv; charset=utf-8"); self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Disposition", "attachment; filename=thimble-findings.csv")
        self.send_header("Content-Length", str(len(payload))); self.end_headers(); self.wfile.write(payload)

    def download_excel(self, params):
        name = Path(params.get("name", [""])[0]).name
        target = (EXCEL_DIR / name).resolve()
        if target.parent != EXCEL_DIR.resolve() or not target.is_file():
            return self.send_error(HTTPStatus.NOT_FOUND)
        payload = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(name)}")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def download_report(self, params):
        name = Path(params.get("name", [""])[0]).name
        target = (REPORT_DIR / name).resolve()
        if target.parent != REPORT_DIR.resolve() or not target.is_file():
            return self.send_error(HTTPStatus.NOT_FOUND)
        payload = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(name)}")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers(); self.wfile.write(payload)

    def serve_static(self, path):
        target = STATIC / ("index.html" if path == "/" else path.lstrip("/"))
        target = target.resolve()
        if STATIC.resolve() not in target.parents and target != STATIC.resolve():
            return self.send_error(HTTPStatus.FORBIDDEN)
        if not target.is_file(): return self.send_error(HTTPStatus.NOT_FOUND)
        payload = target.read_bytes(); self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "application/octet-stream")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Security-Policy", "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; connect-src 'self'; frame-src 'self'")
        self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_header("Content-Length", str(len(payload))); self.end_headers(); self.wfile.write(payload)


def main():
    init_db()
    port = int(os.environ.get("THIMBLE_PORT", "8765"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"指套管检测数据管理系统: http://127.0.0.1:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__": main()
